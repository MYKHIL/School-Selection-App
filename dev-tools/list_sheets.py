from openpyxl import load_workbook
from pathlib import Path

path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
print('Sheet count:', len(wb.sheetnames))
for idx, name in enumerate(wb.sheetnames, start=1):
    print(f'{idx}: {name}')
    if idx <= 20:
        ws = wb[name]
        rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))
        for r in rows:
            print(r)
        print('---')
