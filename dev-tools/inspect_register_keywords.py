from openpyxl import load_workbook
import re
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', read_only=True, data_only=True)
keywords = [r'Appendix', r'Technical Institution', r'Senior High Tech', r'SEN', r'STEM', r'Pilot Private', r'Category A', r'Category B', r'Category C', r'Technical Programmes', r'Public Second Cycle']
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    matches = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
        if row is None:
            continue
        for j, cell in enumerate(row, start=1):
            if cell is None:
                continue
            text = str(cell).strip()
            for kw in keywords:
                if re.search(kw, text, re.IGNORECASE):
                    matches.append((i, j, text))
                    break
        if len(matches) >= 20:
            break
    if matches:
        print('\n===', name, '===')
        for rownum, colnum, text in matches:
            print(rownum, colnum, text)
