from openpyxl import load_workbook
import re
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)
sheet = wb['Sheet37']
rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in sheet.iter_rows(values_only=True)]

for idx,row in enumerate(rows, start=1):
    if '9050101' in row:
        print('FOUND row', idx)
        break


def normalize_cell(value):
    return (value or '').strip()

def is_header_row(row):
    values = [normalize_cell(v) for v in row if normalize_cell(v)]
    numeric_count = sum(1 for v in values if re.match(r'^\d{3,4}$', v))
    return numeric_count >= 5

header_row_idx = None
for idx,row in enumerate(rows):
    if is_header_row(row):
        header_row_idx = idx
        print('header_row_idx', idx+1)
        break

if header_row_idx is None:
    raise SystemExit('No header row')

label_row = rows[header_row_idx-1]
code_row = [normalize_cell(v) for v in rows[header_row_idx]]
program_cols = [ci for ci,v in enumerate(code_row) if re.match(r'^\d{3,4}$', v)]
program_header_map = {}
for ci in program_cols:
    header = normalize_cell(label_row[ci]) or code_row[ci]
    if header and not re.match(r'^(?:S/N|REGION|DISTRICT|INST\.? CODE|INSTITUTION|LOCATION|GENDER|STATUS|CATEGORY|NO\.? OF|APPENDIX)$', header, re.I):
        program_header_map[ci] = header
print('program_header_map keys', list(program_header_map.keys())[:10])

for idx,row in enumerate(rows):
    row_data = [normalize_cell(v) for v in row]
    if '9050101' not in row_data:
        continue
    print('row index', idx+1, row_data[:20])
    print('code index', row_data.index('9050101'))
    programs=[]
    for ci,cell_value in enumerate(row_data):
        if ci == row_data.index('9050101'):
            continue
        if not cell_value:
            continue
        if program_header_map.get(ci) and re.match(r'^(?:X|x|?|?|1|Y|YES)$', cell_value):
            programs.append(program_header_map[ci])
            continue
        if re.match(r'^appendix', cell_value, re.I):
            continue
        if re.search(r'refer to|detailed|programme?s?', cell_value, re.I):
            continue
        if re.match(r'^(?:MIXED|MALES|FEMALES|DAY/BOARDING|DAY|BOARDING|TVET|SHTS|SHS)$', cell_value, re.I):
            continue
        if re.match(r'^\d{1,4}$', cell_value):
            continue
        if re.match(r'^(?:[A-Z]{1,3}\.|\w+\s+INSTITUTE?)$', cell_value, re.I) and ci < row_data.index('9050101'):
            continue
        programs.append(cell_value)
    print('programs', programs)
    break
