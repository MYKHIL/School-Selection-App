import json
import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
json_path = root / 'data' / 'schools_all.json'


def normalize_space(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_label(value):
    value = normalize_space(value)
    value = value.replace(' - ', '-').replace('  ', ' ')
    return value.strip(' ,;:-')


def normalize_school_name(value):
    text = normalize_label(value or '')
    text = re.sub(r"\b(SENIOR HIGH/TECH|SENIOR HIGH TECH|SENIOR HIGH|SR\. HIGH|SNR HIGH|SNR\. HIGH|TECH SCHOOL|TECHNICAL SCHOOL|TECH|SHTS|SHS|STEM|TVET|ACADEMY|INSTITUTE|SCHOOL)\b", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'[^A-Z0-9 ]+', ' ', text.upper())
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def is_marker(value):
    marker = normalize_space(value).upper()
    return marker in {'X', 'XX', '1', 'Y', 'YES', '✔', '✓'}


def parse_shs_shts_appendix(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'S/N' in normalized and 'ALL SENIOR HIGH/ TECH SCHOOLS' in normalized:
            header_index = idx
            break
    if header_index is None:
        return {}
    header_row = rows[header_index]
    subject_headers = [(col_idx, normalize_label(header_row[col_idx])) for col_idx in range(5, len(header_row)) if normalize_label(header_row[col_idx])]
    mapping = {}
    for row in rows[header_index + 1:]:
        if not row:
            continue
        name = normalize_label(row[2]) if len(row) > 2 else ''
        if not name:
            continue
        programs = [label for col_idx, label in subject_headers if col_idx < len(row) and is_marker(row[col_idx])]
        if programs:
            key = normalize_school_name(name) or normalize_label(name).upper()
            mapping.setdefault(key, []).extend(programs)
            mapping[key] = list(dict.fromkeys(mapping[key]))
    return mapping


def parse_tvet_sheet(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        if any(cell and cell.upper() in {'S/N', 'SN'} for cell in row[:6]):
            header_index = idx
            break
    if header_index is None:
        return {}
    header_row = rows[header_index]
    code_row = rows[header_index + 1] if header_index + 1 < len(rows) else []
    mapping = {}
    for row in rows[header_index + 2:]:
        if not row:
            continue
        code = normalize_space(row[3]) if len(row) > 3 else ''
        if not re.fullmatch(r'\d{7}', code):
            continue
        for col_idx in range(7, len(row)):
            if not is_marker(row[col_idx]):
                continue
            label = normalize_label(header_row[col_idx]) if col_idx < len(header_row) else ''
            code_value = normalize_space(code_row[col_idx]) if col_idx < len(code_row) else ''
            if label and not re.fullmatch(r'\d{3,4}', label):
                mapping.setdefault(code, []).append(f'{label} ({code_value})' if re.fullmatch(r'\d{3,4}', code_value) else label)
        if code in mapping:
            mapping[code] = list(dict.fromkeys(mapping[code]))
    return mapping


def merge_programme_maps(target, source):
    for key, values in source.items():
        target.setdefault(key, []).extend(values)
        target[key] = list(dict.fromkeys(target[key]))


def sheet_contains_appendix(ws, appendix_text):
    for row in ws.iter_rows(values_only=True, max_row=8):
        for cell in row:
            if cell and appendix_text in str(cell).upper():
                return True
    return False

wb = load_workbook(xlsx_path, data_only=True, read_only=True)
code_mapping = {}
name_mapping = {}
for sheet in wb.worksheets:
    if sheet_contains_appendix(sheet, 'APPENDIX 1'):
        merge_programme_maps(code_mapping, parse_tvet_sheet(sheet))
    if sheet_contains_appendix(sheet, 'APPENDIX 2'):
        merge_programme_maps(name_mapping, parse_shs_shts_appendix(sheet))
    if sheet_contains_appendix(sheet, 'APPENDIX 4') or 'STEM' in sheet.title.upper():
        merge_programme_maps(code_mapping, parse_tvet_sheet(sheet))

with json_path.open('r', encoding='utf-8') as f:
    schools = json.load(f)

json_by_code = {s.get('code'): s for s in schools}
json_by_name = {normalize_school_name(s.get('name', '')): s for s in schools if s.get('type') in ('SHS','SHTS')}

failures = []
for code, school in json_by_code.items():
    if not school or school.get('type') not in ('SHS','SHTS'):
        continue
    target_keys = []
    if code in code_mapping:
        target_keys.append(f'code:{code}')
    school_key = normalize_school_name(school.get('name', ''))
    if school_key in name_mapping:
        target_keys.append(f'name:{school_key}')
    if code in code_mapping or school_key in name_mapping:
        source_names = []
        if code in code_mapping:
            source_names.extend(code_mapping[code])
        if school_key in name_mapping:
            source_names.extend(name_mapping[school_key])
        source_names = list(dict.fromkeys([normalize_label(n) for n in source_names if normalize_label(n)]))
        current = [normalize_label(n) for n in school.get('programNames', []) if normalize_label(n)]
        missed = [n for n in source_names if n not in current]
        extra = [n for n in current if n not in source_names]
        if missed:
            failures.append((code, school.get('name'), school_key, source_names, current, missed, extra, target_keys))

print('failures', len(failures))
for item in failures[:50]:
    code,name,key,source,current,missed,extra,target_keys = item
    print(code, name)
    print('  source', source)
    print('  current', current)
    print('  missed', missed)
    print('  extra', extra)
    print('  matched', target_keys)
    print()