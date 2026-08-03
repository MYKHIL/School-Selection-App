import openpyxl
from pathlib import Path
path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
for name in wb.sheetnames[:10]:
    ws = wb[name]
    print('\nSheet:', name)
    for r in range(1, 35):
        row = [ws.cell(row=r, column=c).value for c in range(1, 35)]
        if any(cell is not None for cell in row):
            print(r, row)
            if r >= 20:
                # stop after some rows to avoid too much output
                if r > 25:
                    break
