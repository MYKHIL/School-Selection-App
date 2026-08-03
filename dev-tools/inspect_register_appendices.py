from openpyxl import load_workbook

wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', read_only=True, data_only=True)
print('SHEETS:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\n=== SHEET:', name, '===')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True), start=1):
        if any(cell is not None for cell in row):
            print(i, row)
            if i >= 60:
                break
