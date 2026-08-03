from openpyxl import load_workbook
import re

wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)

for name in wb.sheetnames:
    if name == 'Sheet37':
        sheet = wb[name]
        break
else:
    raise SystemExit('Sheet37 not found')

rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in sheet.iter_rows(values_only=True)]

for idx,row in enumerate(rows, start=1):
    if '9050101' in row:
        print('FOUND row', idx)
        print(row[:30])
        break


def normalize_cell(value):
    return (value or '').strip()

def is_header_row(row):
    values = [normalize_cell(v) for v in row if normalize_cell(v)]
    numeric_count = sum(1 for v in values if re.match(r'^\d{3,4}$', v))
    return numeric_count >= 5

header_row_idx = None
for idx,row in enumerate(rows):
    if is_header_row(row):
        header_row_idx = idx
        print('header_row_idx', idx+1)
        break

if header_row_idx is None:
    raise SystemExit('No header row found')

label_row = rows[header_row_idx-1]
code_row = [normalize_cell(v) for v in rows[header_row_idx]]
program_cols = [ci for ci,v in enumerate(code_row) if re.match(r'^\d{3,4}$', v)]
program_header_map = {}
ignore_pattern = re.compile(r'^(?:S/N|REGION|DISTRICT|INST\.? CODE|INSTITUTION|LOCATION|GENDER|STATUS|CATEGORY|NO\.? OF|APPENDIX)$', re.I)
for ci in program_cols:
    header = normalize_cell(label_row[ci]) or code_row[ci]
    if header and not ignore_pattern.match(header):
        program_header_map[ci] = header

print('program_header_map count', len(program_header_map))
for ci in sorted(program_header_map)[:15]:
    print(ci, program_header_map[ci])

marker_regex = re.compile(r'^(?:X|x|?|?|1|Y|YES)$')
for idx,row in enumerate(rows, start=1):
    row_data = [normalize_cell(v) for v in row]
    if '9050101' not in row_data:
        continue
    code_index = row_data.index('9050101')
    programs=[]
    print('row_data length', len(row_data))
    for ci,cell_value in enumerate(row_data):
        if ci == code_index:
            continue
        if not cell_value:
            continue
        if program_header_map.get(ci) and marker_regex.match(cell_value):
            programs.append(program_header_map[ci])
            continue
        if re.match(r'^appendix', cell_value, re.I):
            continue
        if re.search(r'refer to|detailed|programme?s?', cell_value, re.I):
            continue
        if re.match(r'^(?:MIXED|MALES|FEMALES|DAY/BOARDING|DAY|BOARDING|TVET|SHTS|SHS)$', cell_value, re.I):
            continue
        if re.match(r'^\d{1,4}$', cell_value):
            continue
        if re.match(r'^(?:[A-Z]{1,3}\.|\w+\s+INSTITUTE?)$', cell_value, re.I) and ci < code_index:
            continue
        programs.append(cell_value)
    print('programs len', len(programs))
    print(programs[:20])
    break
