import json
from pathlib import Path

from openpyxl import load_workbook

from scripts.export_school_register_json import (
    find_header_row,
    merge_program_map,
    parse_main_register_sheet,
    parse_program_columns_multi_row,
)


def test_ayirebi_school_programs_are_detected():
    root = Path(__file__).resolve().parent.parent
    wb = load_workbook(root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True, read_only=True)
    ws = wb['Sheet26']
    rows = [[str(c).strip() if c is not None else '' for c in row] for row in ws.iter_rows(values_only=True)]
    header_index, _ = find_header_row(ws)
    program_columns, start_idx = parse_program_columns_multi_row(rows, header_index)
    schools = parse_main_register_sheet(ws)

    ayirebi = next(s for s in schools if s['name'].lower() == 'ayirebi senior high')
    assert 'AGRICULTURAL SCIENCE' in ayirebi['programNames']
    assert 'BUSINESS' in ayirebi['programNames']
    assert 'HOME ECONOMICS' in ayirebi['programNames']
    assert 'VISUAL ARTS' in ayirebi['programNames']
    assert 'GENERAL ARTS' in ayirebi['programNames']
    assert program_columns
    assert start_idx <= 8


def test_merge_program_map_preserves_general_and_stem_programs():
    assert merge_program_map(['STEM'], ['GENERAL SCIENCE']) == ['STEM', 'GENERAL SCIENCE']
    assert merge_program_map(['BIO-MEDICAL SCIENCE'], ['GENERAL SCIENCE', 'BUSINESS']) == [
        'BIO-MEDICAL SCIENCE',
        'GENERAL SCIENCE',
        'BUSINESS',
    ]
