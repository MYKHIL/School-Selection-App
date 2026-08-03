from openpyxl import load_workbook
import re

wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)
sheet = wb['Sheet37']

rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in sheet.iter_rows(values_only=True)]
print('rows', len(rows))
for i,row in enumerate(rows[:15], start=1):
    print(i, row[:15])

# find code row

for code in ['9050101','9060200','9050301','9090102']:
    print('--- code', code)
    for idx,row in enumerate(rows, start=1):
        row_data=[str(cell).strip() if cell is not None else '' for cell in row]
        if code in row_data:
            print('found at', idx, row_data[:20])
            code_index=row_data.index(code)
            print('code_index', code_index)
            break
    else:
        print('not found')

# find header via numeric codes
for idx,row in enumerate(rows, start=1):
    values=[str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
    numeric_count=sum(1 for v in values if re.match(r'^\d{3,4}$', v))
    if numeric_count >= 5:
        print('header row', idx, values[:30])
        break

header_idx = 6
label_row = rows[header_idx-1]
code_row = [str(cell).strip() for cell in rows[header_idx]]
program_cols = [ci for ci,v in enumerate(code_row) if re.match(r'^\d{3,4}$', v)]
print('program cols', len(program_cols), program_cols[:20])
for ci in program_cols[:20]:
    print(ci, label_row[ci])

# parse program list for 9050101
code='9050101'
for idx,row in enumerate(rows, start=1):
    row_data=[str(cell).strip() if cell is not None else '' for cell in row]
    if code not in row_data:
        continue
    code_index = row_data.index(code)
    programs=[]
    for ci,cell_value in enumerate(row_data):
        if ci == code_index: continue
        if not cell_value: continue
        if ci in program_cols and re.match(r'^(?:X|x|?|?|1|Y|YES)$', cell_value):
            programs.append(label_row[ci] or code_row[ci])
            continue
        if re.match(r'^appendix', cell_value, re.I): continue
        if re.search(r'refer to|detailed|programme?s?', cell_value, re.I): continue
        if re.match(r'^(?:MIXED|MALES|FEMALES|DAY/BOARDING|DAY|BOARDING|TVET|SHTS|SHS)$', cell_value, re.I): continue
        if re.match(r'^\d{1,4}$', cell_value): continue
        if re.match(r'^(?:[A-Z]{1,3}\.|\w+\s+INSTITUTE?)$', cell_value, re.I) and ci < code_index: continue
        programs.append(cell_value)
    print('programs for', code, len(programs), programs[:20])
