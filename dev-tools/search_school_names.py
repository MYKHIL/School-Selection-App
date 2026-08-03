from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
wb = load_workbook(path, data_only=True, read_only=True)
search_names = [
    'Accra Senior High Asylum Down',
    'Accra Senior High',
    'Ghana Stem Technical School Sewua',
    'KNUST Senior High',
    'Bosomtwe Girls STEM',
    'Abomosu STEM Senior High',
    'Ghana Stem Technical School Sewua',
]
for name in search_names:
    print('SEARCH', name)
    found = []
    for idx, sheet in enumerate(wb.worksheets, start=1):
        title = sheet.title
        for r, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row:
                continue
            for c, cell in enumerate(row, start=1):
                if cell and name.upper() in str(cell).upper():
                    found.append((idx, title, r, c, cell))
    if not found:
        print('  not found')
    else:
        for item in found:
            print(' ', item)
    print()