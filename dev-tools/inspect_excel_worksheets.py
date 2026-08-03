from openpyxl import load_workbook
from pathlib import Path

path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
print('Sheet count:', len(wb.sheetnames))
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    row_count = sum(1 for _ in ws.iter_rows(min_row=1, max_row=100, values_only=True))
    print(f'{idx}: {name} (first 100 rows scanned)')
    headers = []
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if any(cell and isinstance(cell, str) and 'SCHOOL NAME' in cell.upper() for cell in row):
            headers.append((r, row))
    if headers:
        print('  header rows:')
        for r, row in headers:
            print(f'    {r}: {row}')
    first_rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))
    for row in first_rows:
        print('   ', row)
    print('---')
