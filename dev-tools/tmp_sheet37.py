from openpyxl import load_workbook
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)
name = wb.sheetnames[36]
ws = wb[name]
print('sheet', name)
for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
    if i <= 120:
        print(i, [str(cell).strip() if cell is not None else '' for cell in row[:30]])
