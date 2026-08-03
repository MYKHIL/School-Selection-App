from pathlib import Path
from openpyxl import load_workbook
root = Path(r'd:/School Selection App')
path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
wb = load_workbook(path, data_only=True, read_only=True)
print('TOTAL', len(wb.sheetnames))
print('SHEET NAMES 13-20', wb.sheetnames[12:20])
print('SHEET NAMES 37-44', wb.sheetnames[36:44])
for idx in range(12, 20):
    sheet = wb.worksheets[idx]
    print('===', idx+1, sheet.title)
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i > 12:
            break
        print(i, [None if cell is None else str(cell)[:40] for cell in row[:20]])
    print()
print('=== APPENDIX ===')
for idx in range(36, 44):
    sheet = wb.worksheets[idx]
    print('===', idx+1, sheet.title)
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i > 12:
            break
        print(i, [None if cell is None else str(cell)[:40] for cell in row[:20]])
