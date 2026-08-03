import json
import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
json_path = root / 'data' / 'schools_all.json'


def normalize_space(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_label(value):
    value = normalize_space(value)
    value = value.replace(' - ', '-').replace('  ', ' ')
    return value.strip(' ,;:-')


def normalize_school_name(value):
    text = normalize_label(value or '')
    text = re.sub(r"\b(SENIOR HIGH/TECH|SENIOR HIGH TECH|SENIOR HIGH|SR\. HIGH|SNR HIGH|TECH SCHOOL|TECH|SHTS|STEM|TVET|ACADEMY|INSTITUTE|SCHOOL)\b", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'[^A-Z0-9 ]+', ' ', text.upper())
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def is_marker(value):
    marker = normalize_space(value).upper()
    return marker in {'X', 'XX', '1', 'Y', 'YES', '✔', '✓'}


def parse_shs_shts_appendix(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'S/N' in normalized and 'ALL SENIOR HIGH/ TECH SCHOOLS' in normalized:
            header_index = idx
            break
    if header_index is None:
        return {}

    header_row = rows[header_index]
    subject_headers = []
    for col_idx in range(5, len(header_row)):
        label = normalize_label(header_row[col_idx]) if col_idx < len(header_row) else ''
        if label:
            subject_headers.append((col_idx, label))
    mapping = {}
    for row in rows[header_index + 1:]:
        if not row:
            continue
        name = normalize_label(row[2]) if len(row) > 2 else ''
        if not name:
            continue
        programs = []
        for col_idx, label in subject_headers:
            if col_idx < len(row) and is_marker(row[col_idx]):
                programs.append(label)
        if programs:
            mapping[normalize_label(name).upper()] = programs
    return mapping


def parse_tvet_sheet(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        if any(cell and cell.upper() in {'S/N', 'SN'} for cell in row[:6]):
            header_index = idx
            break
    if header_index is None:
        return {}
    header_row = rows[header_index]
    code_row = rows[header_index + 1] if header_index + 1 < len(rows) else []
    mapping = {}
    for row in rows[header_index + 2:]:
        if not row:
            continue
        code = normalize_space(row[3]) if len(row) > 3 else ''
        if not re.fullmatch(r'\d{7}', code):
            continue
        for col_idx in range(7, len(row)):
            if not is_marker(row[col_idx]):
                continue
            label = normalize_label(header_row[col_idx]) if col_idx < len(header_row) else ''
            code_value = normalize_space(code_row[col_idx]) if col_idx < len(code_row) else ''
            if label and not re.fullmatch(r'\d{3,4}', label):
                mapping.setdefault(code, []).append(f'{label} ({code_value})' if re.fullmatch(r'\d{3,4}', code_value) else label)
    return mapping


def merge_programme_maps(target, source):
    for key, values in source.items():
        target.setdefault(key, []).extend(values)
        target[key] = list(dict.fromkeys(target[key]))


def sheet_contains_appendix(ws, appendix_text):
    for row in ws.iter_rows(values_only=True, max_row=8):
        for cell in row:
            if cell and appendix_text in str(cell).upper():
                return True
    return False


wb = load_workbook(xlsx_path, data_only=True, read_only=True)
code_mapping = {}
name_mapping = {}
for sheet in wb.worksheets:
    title = sheet.title
    if sheet_contains_appendix(sheet, 'APPENDIX 1'):
        merge_programme_maps(code_mapping, parse_tvet_sheet(sheet))
    if sheet_contains_appendix(sheet, 'APPENDIX 2'):
        merge_programme_maps(name_mapping, parse_shs_shts_appendix(sheet))
    if sheet_contains_appendix(sheet, 'APPENDIX 4') or 'STEM' in title.upper():
        merge_programme_maps(code_mapping, parse_tvet_sheet(sheet))

with json_path.open('r', encoding='utf-8') as f:
    schools = json.load(f)

json_name_map = {}
for s in schools:
    if s.get('type') in ('SHS', 'SHTS'):
        raw_name = s.get('name', '')
        json_name_map[normalize_label(raw_name).upper()] = raw_name
        json_name_map[normalize_school_name(raw_name)] = raw_name

print('APPENDIX2 name keys', len(name_mapping))
for query in ['MANCEL GIRLS', 'SUNYANI SENIOR HIGH', 'ACCRA SENIOR HIGH', 'KNUST SENIOR HIGH', 'SAINT MARGARET', 'HOLLAND', 'PREMPEH', 'OPUKU WARE']:
    n = normalize_label(query).upper()
    print('QUERY', query)
    print(' exact key exists', n in name_mapping)
    print('  exact json key exists', n in json_name_map)
    print(' json normalized key exists', normalize_school_name(query) in name_mapping)
    matches = [k for k in name_mapping if n in k or k in n]
    print('  substring matches', len(matches), matches[:10])
    print('  sample json names', [k for k in json_name_map.keys() if n in k or k in n][:10])
    print()

# find JSON names that are SHS/SHTS and have only one programmeName
incomplete = [s for s in schools if s.get('type') in ('SHS','SHTS') and len(s.get('programNames',[])) <= 1]
print('incomplete count', len(incomplete))
for s in incomplete[:20]:
    name = s.get('name','')
    print('JSON', s.get('code'), name, normalize_label(name).upper(), normalize_school_name(name))
    if normalize_label(name).upper() in name_mapping:
        print('  exact found in appendix2')
    if normalize_school_name(name) in name_mapping:
        print('  normalized found in appendix2')
    subs = [k for k in name_mapping if normalize_label(name).upper() in k or k in normalize_label(name).upper()]
    if subs:
        print('  substrings', subs[:5])
    print()
