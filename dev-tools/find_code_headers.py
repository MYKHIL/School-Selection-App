import openpyxl
from pathlib import Path
import re
path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('sheets', len(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    matches = []
    for r in range(1, 80):
        row = [ws.cell(row=r, column=c).value for c in range(1, 80)]
        textrow = [str(cell).strip().upper() if cell is not None else '' for cell in row]
        if any('CODE' in cell for cell in textrow) and any('SCHOOL NAME' in cell or 'SCHOOL' == cell for cell in textrow):
            matches.append((r, textrow))
    if matches:
        print('\nSheet:', name)
        for r, textrow in matches:
            print(' header row', r, textrow)
            for rr in range(r, min(r+8, 90)):
                row = [ws.cell(row=rr, column=c).value for c in range(1, 90)]
                if any(cell is not None for cell in row):
                    print(rr, [str(cell).strip() if cell is not None else '' for cell in row])
