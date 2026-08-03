import json
import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
json_path = root / 'data' / 'schools_all.json'


def normalize_space(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_label(value):
    value = normalize_space(value)
    value = value.replace(' - ', '-').replace('  ', ' ')
    return value.strip(' ,;:-')


def normalize_school_name(value):
    text = normalize_label(value or '')
    text = re.sub(r"\b(SENIOR HIGH/TECH|SENIOR HIGH TECH|SENIOR HIGH|SR\. HIGH|SNR HIGH|SNR\. HIGH|TECH SCHOOL|TECHNICAL SCHOOL|TECH|SHTS|SHS|STEM|TVET|ACADEMY|INSTITUTE|SCHOOL)\b", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'[^A-Z0-9 ]+', ' ', text.upper())
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def search_workbook(names):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    hits = {}
    for sheet in wb.worksheets:
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_text = ' '.join(str(cell or '') for cell in row).upper()
            for key, school_name, normalized in names:
                if key in row_text or normalized in row_text:
                    hits.setdefault(key, []).append((sheet.title, row_idx, row))
    return hits


with json_path.open('r', encoding='utf-8') as f:
    schools = json.load(f)

candidates = [s for s in schools if s.get('type') in ('SHS','SHTS') and len(s.get('programNames',[])) <= 1]
print('candidate count', len(candidates))
selected = candidates[:50]
search_list = []
for s in selected:
    name = s.get('name','')
    key = normalize_label(name).upper()
    normalized = normalize_school_name(name)
    search_list.append((key, name, normalized))

hits = search_workbook(search_list)
for key, name, normalized in search_list:
    print('SEARCH', key, normalized, 'ORIGINAL', name)
    if key in hits:
        print('  exact matches', len(hits[key]))
        for sheet, row_idx, row in hits[key][:3]:
            print('   ', sheet, row_idx, row)
    elif normalized in hits:
        print('  normalized matches', len(hits[normalized]))
        for sheet, row_idx, row in hits[normalized][:3]:
            print('   ', sheet, row_idx, row)
    else:
        print('  NO MATCH')
    print()