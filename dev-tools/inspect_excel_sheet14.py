import openpyxl
from pathlib import Path
path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('sheetcount', len(wb.sheetnames))
for name in wb.sheetnames:
    if name not in ('Sheet14', 'Sheet15', 'Sheet16', 'Sheet17', 'Sheet18'):
        continue
    ws = wb[name]
    print('\n=== Sheet', name, '===')
    for r in range(1, 40):
        row = [ws.cell(row=r, column=c).value for c in range(1, 50)]
        if any(cell is not None for cell in row):
            print(r, [str(cell).strip() if cell is not None else '' for cell in row])
