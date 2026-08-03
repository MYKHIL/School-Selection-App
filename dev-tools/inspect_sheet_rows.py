from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
wb = load_workbook(path, data_only=True, read_only=True)
for idx in [43, 47, 52]:
    sheet = wb.worksheets[idx - 1]
    print('===', idx, sheet.title)
    for r, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if r > 30:
            break
        print(r, [None if cell is None else str(cell) for cell in row[:25]])
    print()
