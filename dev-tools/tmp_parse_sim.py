from openpyxl import load_workbook
import re
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)
sheet = wb['Sheet37']
rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in sheet.iter_rows(values_only=True)]

for i,row in enumerate(rows, start=1):
    if '9050101' in row:
        print('ROW', i, row)
        break


def normalize_cell(value):
    return (value or '').strip()

def is_header_row(row):
    values = [normalize_cell(v) for v in row if normalize_cell(v)]
    numeric_count = sum(1 for v in values if re.match(r'^\d{3,4}$', v))
    return numeric_count >= 5

header_idx = None
for idx,row in enumerate(rows):
    if is_header_row(row):
        header_idx = idx
        print('HEADER idx', idx+1, row)
        break

label_row = rows[header_idx-1] if header_idx is not None and header_idx-1>=0 else rows[header_idx]
print('LABEL row index', header_idx, label_row)

code_row = [normalize_cell(v) for v in rows[header_idx]]
program_cols = [ci for ci,v in enumerate(code_row) if re.match(r'^\d{3,4}$', v)]
print('program_cols', program_cols)
print('code_row values', [code_row[ci] for ci in program_cols])

program_header_map = {}
for ci in program_cols:
    header = normalize_cell(label_row[ci]) or code_row[ci]
    if header and not re.match(r'^(?:S/N|REGION|DISTRICT|INST\.? CODE|INSTITUTION|LOCATION|GENDER|STATUS|CATEGORY|NO\.? OF|APPENDIX)$', header, re.I):
        program_header_map[ci] = header
print('program_header_map len', len(program_header_map))
for ci, hv in program_header_map.items():
    print(ci, hv)

# now extract programs for sample row
for row in rows:
    if '9050101' in row:
        codeIndex = row.index('9050101')
        programs=[]
        for ci,val in enumerate(row):
            if ci==codeIndex: continue
            cellValue = normalize_cell(val)
            if not cellValue: continue
            if program_header_map.get(ci) and re.match(r'^(?:X|x|?|?|1|Y|YES)$', cellValue):
                programs.append(program_header_map[ci])
                continue
            if re.match(r'^appendix', cellValue, re.I): continue
            if re.search(r'refer to|detailed|programme?s?', cellValue, re.I): continue
            if re.match(r'^(?:MIXED|MALES|FEMALES|DAY/BOARDING|DAY|BOARDING|TVET|SHTS|SHS)$', cellValue, re.I): continue
            if re.match(r'^\d{1,4}$', cellValue): continue
            if re.match(r'^(?:[A-Z]{1,3}\.|\w+\s+INSTITUTE?)$', cellValue, re.I) and ci < codeIndex: continue
            programs.append(cellValue)
        print('programs', programs)
        break
