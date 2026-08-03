import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
json_path = root / 'data' / 'schools_all.json'
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'


def normalize_space(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_label(value):
    value = normalize_space(value)
    value = value.replace(' - ', '-').replace('  ', ' ')
    return value.strip(' ,;:-')


def normalize_school_name(value):
    text = normalize_label(value or '')
    text = re.sub(r"\b(SENIOR HIGH/TECH|SENIOR HIGH TECH|SENIOR HIGH|SR\. HIGH|SNR HIGH|SNR\. HIGH|TECH SCHOOL|TECHNICAL SCHOOL|TECH|SHTS|SHS|STEM|TVET|ACADEMY|INSTITUTE|SCHOOL)\b", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r"(\bST\.?\b|\bSN\.?\b)", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'[^A-Z0-9 ]+', ' ', text.upper())
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def is_marker(value):
    marker = normalize_space(value).upper()
    return marker in {'X', 'XX', '1', 'Y', 'YES', '✔', '✓'}


def decorate_programme_name(label, code=None):
    label = normalize_label(label)
    if not label:
        return ''
    code = normalize_space(code)
    if re.fullmatch(r'\d{3,4}', code):
        if re.search(r'\(\d{3,4}\)\s*$', label):
            return label
        return f'{label} ({code})'
    return label


def canonical_progs_for_program_names(names):
    if not names:
        return []
    mapping = {
        'GENERAL SCIENCE': 'GEN. SCI',
        'GENERAL ARTS': 'GEN. ARTS',
        'LANGUAGES': 'GEN. ARTS',
        'BUSINESS': 'BUS',
        'HOME ECONOMICS': 'HOM. ECON.',
        'VISUAL ARTS': 'VIS. ARTS',
        'TECHNICAL PROGRAMMES': 'TECH',
        'TECHNICAL': 'TECH',
        'TVET': 'TECH',
        'STEM': 'STEM',
        'BIO-MEDICAL SCIENCE': 'STEM',
        'ENGINEERING SCIENCE': 'STEM',
        'AVIATION & AEROSPACE ENGINEERING': 'STEM',
        'COMPUTING': 'STEM',
        'ROBOTICS': 'STEM',
        'AGRICULTURAL SCIENCE': 'AGRIC',
        'MANUFACTURING ENGINEERING': 'STEM'
    }
    canonical = []
    for value in names:
        normalized = normalize_label(value).upper()
        if not normalized:
            continue
        if normalized in mapping:
            canonical.append(mapping[normalized])
            continue
        if re.search(r'\b(STEM|BIO|ENGINEERING|COMPUTING|ROBOTICS|AEROSPACE)\b', normalized):
            canonical.append('STEM')
            continue
        if re.search(r'\b(TECH|TVET|VOCATIONAL|TRADE|TRADES|CONSTRUCTION|AUTOMOTIVE|MECHANICAL|ELECTRICAL|HOSPITALITY|COSMETOLOGY)\b', normalized):
            canonical.append('TECH')
            continue
        if re.search(r'\b(AGRIC|AGRO|CROP|ANIMAL)\b', normalized):
            canonical.append('AGRIC')
            continue
        if re.search(r'\b(BUS|ACCOUNTING|FINANCE|MARKETING)\b', normalized):
            canonical.append('BUS')
            continue
        if re.search(r'\b(HOME\s*ECON|HOM\. ECON|FOOD|TEXTILES|CATERING)\b', normalized):
            canonical.append('HOM. ECON.')
            continue
        if re.search(r'\b(VIS|VISUAL)\s*ARTS?\b', normalized):
            canonical.append('VIS. ARTS')
            continue
        if re.search(r'\b(GEN(?:ERAL)?\s*ARTS?)\b', normalized):
            canonical.append('GEN. ARTS')
            continue
        if re.search(r'\b(GEN(?:ERAL)?\s*SCI(?:ENCE)?)\b', normalized):
            canonical.append('GEN. SCI')
            continue
    return list(dict.fromkeys(canonical))


def sheet_contains_appendix(ws, appendix_text):
    for row in ws.iter_rows(values_only=True, max_row=8):
        for cell in row:
            if cell and appendix_text in str(cell).upper():
                return True
    return False


def parse_tvet_sheet(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        if any(cell.upper() in {'S/N', 'SN'} for cell in row[:6]):
            header_index = idx
            break
    if header_index is None:
        return {}

    header_row = rows[header_index]
    code_row = rows[header_index + 1] if header_index + 1 < len(rows) else []
    mapping = {}
    for row in rows[header_index + 2:]:
        if not row:
            continue
        code = normalize_space(row[3]) if len(row) > 3 else ''
        if not re.fullmatch(r'\d{7}', code):
            continue
        for col_idx in range(7, len(row)):
            if not is_marker(row[col_idx]):
                continue
            label = normalize_label(header_row[col_idx]) if col_idx < len(header_row) else ''
            code_value = normalize_space(code_row[col_idx]) if col_idx < len(code_row) else ''
            if label and not re.fullmatch(r'\d{3,4}', label):
                mapping.setdefault(code, []).append(decorate_programme_name(label, code_value))
    return mapping


def parse_shs_shts_appendix(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'S/N' in normalized and 'ALL SENIOR HIGH/ TECH SCHOOLS' in normalized:
            header_index = idx
            break
    if header_index is None:
        return {}

    header_row = rows[header_index]
    subject_headers = []
    for col_idx in range(5, len(header_row)):
        label = normalize_label(header_row[col_idx]) if col_idx < len(header_row) else ''
        if label:
            subject_headers.append((col_idx, label))
    if not subject_headers:
        return {}

    mapping = {}
    for row in rows[header_index + 1:]:
        if not row:
            continue
        name = normalize_label(row[2]) if len(row) > 2 else ''
        if not name:
            continue
        programs = []
        for col_idx, label in subject_headers:
            if col_idx < len(row) and is_marker(row[col_idx]):
                programs.append(label)
        if programs:
            key = normalize_school_name(name)
            if not key:
                key = normalize_label(name).upper()
            mapping.setdefault(key, []).extend(programs)
            mapping[key] = list(dict.fromkeys(mapping[key]))
    return mapping


def parse_stem_sheet(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    mapping = {}
    label_by_col = {
        7: 'BIO-MEDICAL SCIENCE',
        8: 'ENGINEERING SCIENCE',
        9: 'AVIATION & AEROSPACE ENGINEERING',
        10: 'COMPUTING',
        11: 'ROBOTICS',
        12: 'AGRICULTURAL SCIENCE',
        13: 'MANUFACTURING ENGINEERING',
    }
    for row in rows[10:]:
        if not row:
            continue
        code = normalize_space(row[3]) if len(row) > 3 else ''
        if not code:
            continue
        code = re.sub(r'\D', '', code)
        if len(code) == 5:
            code = '00' + code
        if len(code) == 6:
            code = '0' + code
        if len(code) != 7:
            continue
        for col_idx, label in label_by_col.items():
            if col_idx >= len(row):
                continue
            if is_marker(row[col_idx]):
                mapping.setdefault(code, []).append(label)
    return mapping


def parse_shs_sheet(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'CODE' in normalized and 'SCHOOL NAME' in normalized and 'GENDER' in normalized:
            header_index = idx
            break
    if header_index is None or header_index < 1:
        return {}

    group_row = rows[header_index - 1]
    header_row = rows[header_index]
    programme_columns = []
    for col_idx in range(7, len(header_row)):
        sublabel = normalize_label(header_row[col_idx]) if col_idx < len(header_row) else ''
        grouplabel = normalize_label(group_row[col_idx]) if col_idx < len(group_row) else ''
        if grouplabel and sublabel:
            label = f'{grouplabel} {sublabel}'.strip()
        else:
            label = sublabel or grouplabel
        if label:
            programme_columns.append((col_idx, label))

    mapping = {}
    for row in rows[header_index + 1:]:
        if not row or len(row) <= 3:
            continue
        code = normalize_space(row[3])
        code = re.sub(r'\D', '', code)
        if len(code) == 5:
            code = '00' + code
        if len(code) == 6:
            code = '0' + code
        if len(code) != 7:
            continue
        programmes = []
        for col_idx, label in programme_columns:
            if col_idx < len(row) and is_marker(row[col_idx]):
                programmes.append(label)
        if programmes:
            mapping.setdefault(code, []).extend(programmes)
            mapping[code] = list(dict.fromkeys(mapping[code]))
    return mapping


def merge_programme_maps(target, source):
    for key, values in source.items():
        if not values:
            continue
        target.setdefault(key, []).extend(values)
        target[key] = list(dict.fromkeys(target[key]))


def load_workbook_programmes():
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    code_mapping = {}
    name_mapping = {}

    for sheet in wb.worksheets:
        if sheet_contains_appendix(sheet, 'APPENDIX 1'):
            merge_programme_maps(code_mapping, parse_tvet_sheet(sheet))
        if sheet_contains_appendix(sheet, 'APPENDIX 2'):
            merge_programme_maps(name_mapping, parse_shs_shts_appendix(sheet))
        if sheet_contains_appendix(sheet, 'APPENDIX 4') or 'STEM' in sheet.title.upper():
            merge_programme_maps(code_mapping, parse_stem_sheet(sheet))

    return code_mapping, name_mapping


with json_path.open('r', encoding='utf-8') as f:
    schools = json.load(f)

program_names_by_code, program_names_by_name = load_workbook_programmes()

for school in schools:
    code = str(school.get('code', '')).strip()
    existing_names = [normalize_label(value) for value in school.get('programNames', []) if normalize_label(value)]
    names = list(dict.fromkeys(existing_names))
    if code in program_names_by_code:
        names.extend([normalize_label(value) for value in program_names_by_code[code] if normalize_label(value)])
    school_name_key = normalize_school_name(school.get('name', ''))
    fallback_label_key = normalize_label(school.get('name', '')).upper()
    if school_name_key and school_name_key in program_names_by_name:
        names.extend(program_names_by_name[school_name_key])
    elif fallback_label_key in program_names_by_name:
        names.extend(program_names_by_name[fallback_label_key])
    else:
        best_key = None
        best_score = 0
        school_tokens = set(school_name_key.split()) if school_name_key else set(fallback_label_key.split())
        for key in program_names_by_name:
            key_tokens = set(key.split())
            common = school_tokens & key_tokens
            score = len(common)
            if score > best_score:
                best_score = score
                best_key = key
        if best_key and best_score >= 2:
            names.extend(program_names_by_name[best_key])
    names = list(dict.fromkeys(names))

    if not names:
        names = {
            'TVET': ['TECHNICAL PROGRAMMES'],
            'STEM': ['STEM'],
            'SHTS': ['GENERAL SCIENCE'],
            'SHS': ['GENERAL SCIENCE'],
        }.get(school.get('type'), ['GENERAL SCIENCE'])

    school['programNames'] = names

    existing_progs = [normalize_label(value) for value in school.get('progs', []) if normalize_label(value)]
    broad_progs = canonical_progs_for_program_names(names)
    if school.get('type') == 'TVET':
        progs = existing_progs or ['TECH']
    elif school.get('type') == 'STEM':
        progs = ['STEM']
        if existing_progs:
            progs = existing_progs + progs
        if broad_progs:
            progs.extend([prog for prog in broad_progs if prog not in progs])
    else:
        if existing_progs:
            progs = existing_progs
            if broad_progs:
                progs.extend([prog for prog in broad_progs if prog not in progs])
        else:
            progs = broad_progs or ['GEN. SCI']

    school['progs'] = list(dict.fromkeys(progs))

with json_path.open('w', encoding='utf-8') as f:
    json.dump(schools, f, ensure_ascii=False, indent=2)

schools_by_region = defaultdict(list)
for school in schools:
    schools_by_region[school.get('region', 'Unknown')].append(school)
with (root / 'data' / 'schools_by_region.json').open('w', encoding='utf-8') as f:
    json.dump(dict(sorted(schools_by_region.items())), f, ensure_ascii=False, indent=2)

print(f'Updated programme names for {len(program_names_by_code)} Excel school codes; records written: {len(schools)}')
