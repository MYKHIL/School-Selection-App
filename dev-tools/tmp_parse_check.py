from openpyxl import load_workbook
import re
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)

appendix_sheets = {}
appendix_candidates = []
for name in wb.sheetnames:
    sheet = wb[name]
    uname = name.upper()
    m = re.search(r'APPENDIX\s*(\d+)', uname)
    if m:
        appendix_sheets[m.group(1)] = sheet
    elif 'APPENDIX' in uname:
        m2 = re.search(r'(\d+)', uname)
        if m2:
            appendix_sheets[m2.group(1)] = sheet
        else:
            appendix_candidates.append(sheet)
    else:
        rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in sheet.iter_rows(min_row=1, max_row=40, values_only=True)]
        sample_text = '\n'.join([' '.join(r) for r in rows]).upper()
        if re.search(r'APPENDIX|DETAILED TECHNICAL|TECHNICAL PROGRAMME|PROGRAMME|PROGRAMS|PROGRAMMES|APPENDIX\s*\d+', sample_text):
            m3 = re.search(r'APPENDIX\s*(\d+)', sample_text)
            if m3:
                appendix_sheets[m3.group(1)] = sheet
            else:
                appendix_candidates.append(sheet)
        else:
            code_count = len(re.findall(r'\b\d{7}\b', sample_text))
            if code_count >= 3:
                appendix_candidates.append(sheet)

print('appendix sheets', list(appendix_sheets.keys()), 'candidates', len(appendix_candidates))


def parse_programs_from_appendix(appendix_sheet, school_code):
    if appendix_sheet is None:
        return []
    rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in appendix_sheet.iter_rows(values_only=True)]
    if not rows:
        return []

    def normalize_cell(value):
        return (value or '').strip()

    def is_header_row(row):
        values = [normalize_cell(v) for v in row if normalize_cell(v)]
        numeric_count = sum(1 for v in values if re.match(r'^\d{3,4}$', v))
        return numeric_count >= 5

    header_row_idx = None
    for idx,row in enumerate(rows):
        if is_header_row(row):
            header_row_idx = idx
            break
    if header_row_idx is None:
        return []

    label_row = rows[header_row_idx-1] if header_row_idx-1 >=0 else rows[header_row_idx]
    code_row = [normalize_cell(v) for v in rows[header_row_idx]]
    program_cols = [ci for ci,v in enumerate(code_row) if re.match(r'^\d{3,4}$', v)]
    program_header_map = {}
    for ci in program_cols:
        header = normalize_cell(label_row[ci]) or code_row[ci]
        if header and not re.match(r'^(?:S/N|REGION|DISTRICT|INST\.? CODE|INSTITUTION|LOCATION|GENDER|STATUS|CATEGORY|NO\.? OF|APPENDIX)$', header, re.I):
            program_header_map[ci] = header

    for row in rows:
        row_data=[normalize_cell(v) for v in row]
        if school_code not in row_data:
            continue
        code_index = row_data.index(school_code)
        programs=[]
        for ci,cell_value in enumerate(row_data):
            if ci == code_index:
                continue
            if not cell_value:
                continue
            if program_header_map.get(ci) and re.match(r'^(?:X|x|?|?|1|Y|YES)$', cell_value):
                programs.append(program_header_map[ci])
                continue
            if re.match(r'^appendix', cell_value, re.I):
                continue
            if re.search(r'refer to|detailed|programme?s?', cell_value, re.I):
                continue
            if re.match(r'^(?:MIXED|MALES|FEMALES|DAY/BOARDING|DAY|BOARDING|TVET|SHTS|SHS)$', cell_value, re.I):
                continue
            if re.match(r'^\d{1,4}$', cell_value):
                continue
            if re.match(r'^(?:[A-Z]{1,3}\.|\w+\s+INSTITUTE?)$', cell_value, re.I) and ci < code_index:
                continue
            programs.append(cell_value)
        return list(dict.fromkeys(programs))
    return []


def find_programs_for_code(appendix_num, school_code):
    if appendix_num:
        sheet = appendix_sheets.get(appendix_num)
        res = parse_programs_from_appendix(sheet, school_code)
        if res:
            return res
    for key,sheet in appendix_sheets.items():
        res = parse_programs_from_appendix(sheet, school_code)
        if res:
            return res
    for sheet in appendix_candidates:
        res = parse_programs_from_appendix(sheet, school_code)
        if res:
            return res
    return []

for code in ['9050101','9060200','9050301','9090102']:
    print('CODE', code)
    print('explicit 1', find_programs_for_code('1', code))
    print('search all', find_programs_for_code(None, code))
