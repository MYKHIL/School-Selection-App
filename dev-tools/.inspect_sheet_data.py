from pathlib import Path
import openpyxl
import re

path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
print('Workbook exists:', path.exists())
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\n===', name, 'max_row', ws.max_row, 'max_column', ws.max_column)
    rows = list(ws.iter_rows(values_only=True))
    nonempty_rows = [row for row in rows if any(cell is not None for cell in row)]
    print('nonempty rows', len(nonempty_rows))
    print('sample first 10 nonempty rows:')
    for r, row in enumerate(nonempty_rows[:10], 1):
        print(r, row)
    codes = []
    for r, row in enumerate(nonempty_rows, 1):
        for c, cell in enumerate(row, 1):
            if cell is None:
                continue
            text = str(cell).strip()
            if re.search(r'\b\d{7}\b', text) or re.fullmatch(r'\d{3,7}(?:\.0+)?', text):
                codes.append((r, c, text, row))
                break
    print('code-containing rows', len(codes), 'sample', codes[:10])
