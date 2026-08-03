from openpyxl import load_workbook
import re
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)
sheet = wb['Sheet37']
rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in sheet.iter_rows(values_only=True)]

marker_regex = re.compile(r'^(?:X|x|?|?|1|Y|YES)$')
for idx,row in enumerate(rows, start=1):
    if '9050101' in row:
        print('FOUND row', idx)
        code_index = row.index('9050101')
        print('code_index', code_index)
        print(row[:30])
        break

print('processing header rows')
for idx,row in enumerate(rows, start=1):
    values=[str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
    if len([v for v in values if re.match(r'^\d{3,4}$', v)]) >= 5:
        print('header row', idx, values[:40])
        header_idx=idx-1
        break
print('header_idx', header_idx)
label_row=rows[header_idx-1]
code_row=[str(cell).strip() for cell in rows[header_idx]]
program_cols=[ci for ci,v in enumerate(code_row) if re.match(r'^\d{3,4}$', v)]
print('program_cols len', len(program_cols), program_cols[:20])
for ci in program_cols[:15]:
    print(ci, label_row[ci])

row_data=[str(cell).strip() if cell is not None else '' for cell in rows[41]]
print('row length', len(row_data))
programs=[]
for ci,cell_value in enumerate(row_data):
    if ci == code_index: continue
    if not cell_value: continue
    if ci in program_cols and marker_regex.fullmatch(cell_value):
        programs.append(label_row[ci] or code_row[ci])
        continue
    if re.match(r'^appendix', cell_value, re.I): continue
    if re.search(r'refer to|detailed|programme?s?', cell_value, re.I): continue
    if re.match(r'^(?:MIXED|MALES|FEMALES|DAY/BOARDING|DAY|BOARDING|TVET|SHTS|SHS)$', cell_value, re.I): continue
    if re.match(r'^\d{1,4}$', cell_value): continue
    if re.match(r'^(?:[A-Z]{1,3}\.|\w+\s+INSTITUTE?)$', cell_value, re.I) and ci < code_index: continue
    programs.append(cell_value)
print('programs len', len(programs))
print(programs)
