import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

# Define project root and default output paths
root = Path(__file__).resolve().parent.parent
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
output_dir = root / 'data'
output_dir.mkdir(exist_ok=True)

# Valid marker strings indicating program offering in Excel check columns
MARKER_VALUES = {'X', 'XX', '1', '1.0', 'Y', 'YES', '✔', '✓', 'TRUE'}

# Ghana's 16 administrative regions (composite regions first to prevent false substring matches)
GHANA_REGIONS = [
    'Greater Accra', 'Western North', 'Bono East', 'North East',
    'Upper West', 'Upper East', 'Eastern', 'Central', 'Western',
    'Ashanti', 'Volta', 'Northern', 'Bono', 'Ahafo', 'Oti', 'Savannah'
]

# Comprehensive GES 3-digit School Code prefix mapping (001-016 and 901-916)
PREFIX_REGIONS = {
    '001': 'Greater Accra', '901': 'Greater Accra',
    '002': 'Eastern',       '902': 'Eastern',
    '003': 'Central',       '903': 'Central',
    '004': 'Western',       '904': 'Western',
    '005': 'Ashanti',       '905': 'Ashanti',
    '006': 'Volta',         '906': 'Volta',
    '007': 'Northern',      '907': 'Northern',
    '008': 'Upper West',    '908': 'Upper West',
    '009': 'Upper East',    '909': 'Upper East',
    '010': 'Bono',          '910': 'Bono',
    '011': 'Bono East',     '911': 'Bono East',
    '012': 'Ahafo',         '912': 'Ahafo',
    '013': 'Oti',           '913': 'Oti',
    '014': 'Savannah',      '914': 'Savannah',
    '015': 'North East',    '915': 'North East',
    '016': 'Western North', '916': 'Western North',
}

CATEGORY_KEYWORDS = {
    'CATEGORY A': 'A',
    'CATEGORY B': 'B',
    'CATEGORY C': 'C',
    'CATEGORY D': 'D',
}

# Metadata header titles that should never be treated as program names
METADATA_IGNORE_HEADERS = {
    'S/N', 'SN', 'NO', 'NO.', 'S/NO', 'REGION', 'DISTRICT', 'CODE', 'SCHOOL', 'SCHOOL NAME',
    'NAME OF SCHOOL', 'INSTITUTION', 'LOCATION', 'TOWN', 'CITY', 'GENDER', 'SEX', 'BOYS/GIRLS',
    'STATUS', 'TYPE', 'SCHOOL TYPE', 'CATEGORY', 'CAT', 'REMARKS', 'TOTAL', 'SUBTOTAL', 'GRAND TOTAL',
    'NO. OF', 'NO OF', 'CLASS', 'LEVEL', 'AGE', 'ADDRESS', 'POSTAL', 'INST CODE', 'INST. CODE', 'POSTAL ADDRESS'
}

# Standard explicit program names and their aliases
STANDARD_PROGRAM_ALIASES = [
    ('AGRICULTURAL SCIENCE', ['AGRICULTURAL SCIENCE', 'AGRIC SCIENCE', 'AGRIC', 'AGRICULTURE', 'GEN AGRIC']),
    ('BUSINESS', ['BUSINESS', 'BUS', 'COMMERCIAL', 'SECRETARIAL']),
    ('HOME ECONOMICS', ['HOME ECONOMICS', 'HOM. ECON.', 'HOME ECON', 'HOM ECON', 'CATERING', 'HOSPITALITY']),
    ('VISUAL ARTS', ['VISUAL ARTS', 'VIS. ARTS', 'VISUAL ART', 'VIS ARTS', 'GRAPHIC DESIGN']),
    ('GENERAL ARTS', ['GENERAL ARTS', 'GEN. ARTS', 'GEN ARTS', 'LANGUAGES', 'ARTS']),
    ('GENERAL SCIENCE', ['GENERAL SCIENCE', 'GEN. SCI', 'GEN SCI', 'SCIENCE']),
    ('TECHNICAL', ['TECHNICAL PROGRAMMES', 'TECHNICAL', 'TECH PROGRAMMES', 'TECH', 'TVET']),
    ('STEM', ['STEM', 'ROBOTICS', 'ENGINEERING', 'COMPUTING', 'AEROSPACE', 'BIO-MEDICAL']),
]

def normalize_space(value):
    """Normalize consecutive whitespaces into a single space."""
    return re.sub(r'\s+', ' ', str(value or '')).strip()

def normalize_label(value):
    """Clean label strings by stripping trailing punctuation and extra spaces."""
    value = normalize_space(value)
    value = value.replace(' - ', '-').replace('  ', ' ')
    return value.strip(' ,;:-')

def normalize_code(value):
    """Normalize Excel school codes, cleaning float artifacts and padding numeric values to 7 digits."""
    val_str = str(value or '').strip()
    if not val_str or val_str.upper() in ('NONE', 'N/A', 'CODE', 'INST CODE', 'SCHOOL CODE', 'S/N', 'SN', 'NULL'):
        return None
    
    # Strip Excel float conversion artifacts (e.g. '10101.0' -> '10101')
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    
    digits = re.sub(r'\D', '', val_str)
    if not digits:
        # Alphanumeric codes (e.g., TVET/STEM codes)
        clean = re.sub(r'[^A-Z0-9]', '', val_str.upper())
        return clean if len(clean) >= 3 else None
    
    # Standard GES 7-digit school codes padding
    if len(digits) <= 7:
        padded = digits.zfill(7)
        prefix = padded[:3]
        # Valid GES school codes must have a recognized region prefix (001-016 or 901-916)
        if prefix in PREFIX_REGIONS:
            return padded
        # Ignore serial numbers (S/N 1, 2, 153) that pad to '0000xxx'
        if len(digits) >= 5:
            return padded
        return None
    return digits

def is_marker(value):
    """Check if cell value represents an active selection marker."""
    val_str = normalize_space(value).upper()
    return val_str in MARKER_VALUES or val_str == '1' or val_str == '1.0'

def normalize_school_name(value):
    """Normalize school name for reliable cross-sheet lookups."""
    text = normalize_label(value or '')
    text = re.sub(r"\b(SENIOR HIGH/TECH|SENIOR HIGH TECH|SENIOR HIGH|SR\. HIGH|SNR HIGH|SNR\. HIGH|TECH SCHOOL|TECHNICAL SCHOOL|TECH|SHTS|SHS|TVET|ACADEMY|INSTITUTE|SCHOOL|COMMUNITY|DAY|BOARDING)\b", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r"(\bST\.?\b|\bSN\.?\b)", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'[^A-Z0-9 ]+', ' ', text.upper())
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def infer_region_from_code(code):
    """Infer Ghana administrative region from 3-digit GES school code prefix."""
    if not code or len(code) < 3:
        return None
    return PREFIX_REGIONS.get(code[:3])

def infer_region_from_text(text):
    """Detect Ghana region name or major city/town name within text string."""
    if not text:
        return None
    text_upper = str(text).upper()
    region_aliases = {
        'GT. ACCRA': 'Greater Accra',
        'GT ACCRA': 'Greater Accra',
        'GREAT ACCRA': 'Greater Accra',
        'GREATER ACCRA': 'Greater Accra',
        'UPPER EAST': 'Upper East',
        'UPPER WEST': 'Upper West',
        'BONO EAST': 'Bono East',
        'WESTERN NORTH': 'Western North',
    }
    for alias, region in region_aliases.items():
        if alias in text_upper:
            return region
    for region in GHANA_REGIONS:
        if region.upper() in text_upper:
            return region
            
    # City to region mapping for fallback region resolution
    city_map = {
        'ACCRA': 'Greater Accra', 'KUMASI': 'Ashanti', 'TAMALE': 'Northern',
        'CAPE COAST': 'Central', 'SEKONDI': 'Western', 'TAKORADI': 'Western',
        'SUNYANI': 'Bono', 'KOFORIDUA': 'Eastern', 'HO': 'Volta',
        'WA': 'Upper West', 'BOLGATANGA': 'Upper East', 'TECHIMAN': 'Bono East',
        'DORMAA': 'Bono', 'BEREKUM': 'Bono', 'NKAWKAW': 'Eastern',
        'TARKWA': 'Western', 'GOASO': 'Ahafo', 'DAMONGO': 'Savannah',
        'WALEWALE': 'North East', 'NALERIGU': 'North East',
        'SEFWI WIAWSO': 'Western North', 'KADJEBI': 'Oti', 'DAMBAI': 'Oti'
    }
    for city, reg in city_map.items():
        if city in text_upper:
            return reg
    return None

def infer_type_from_value(value):
    """Infer institution type (SHS, SHTS, TVET, STEM) from value text."""
    if not value:
        return None
    text = normalize_label(value).upper()
    if any(k in text for k in ['TVET', 'TECHNICAL INSTITUTE', 'VOCATIONAL', 'COMMERCIAL INSTITUTE', 'TECHNICAL COLL']):
        return 'TVET'
    if 'STEM' in text:
        return 'STEM'
    if any(k in text for k in ['SHTS', 'SHS/TECH', 'SENIOR HIGH/TECH', 'SENIOR HIGH TECH', 'TECHNICAL/COMMERCIAL']):
        return 'SHTS'
    if any(k in text for k in ['SHS', 'SENIOR HIGH', 'HIGH SCHOOL', 'ACADEMY', 'SECD', 'SECONDARY']):
        return 'SHS'
    return None

def infer_status_from_row(row):
    """Infer residency status (Day, Boarding, Day/Boarding) from row end columns."""
    tail = ' '.join(str(cell or '') for cell in row[-4:]).upper()
    if 'DAY/BOARDING' in tail or 'DAY & BOARDING' in tail or ('DAY' in tail and 'BOARDING' in tail):
        return 'Day/Boarding'
    if 'BOARDING' in tail or 'BDG' in tail:
        return 'Boarding'
    if 'DAY' in tail:
        return 'Day'
    return 'Day/Boarding'

def merge_header_rows(primary, secondary):
    merged = []
    max_len = max(len(primary), len(secondary))
    for i in range(max_len):
        primary_val = primary[i] if i < len(primary) else None
        secondary_val = secondary[i] if i < len(secondary) else None
        merged.append(normalize_label(primary_val) or normalize_label(secondary_val) or '')
    return merged


def find_header_row(ws):
    """Find primary header row index and normalized cell values in worksheet."""
    rows = list(ws.iter_rows(values_only=True, max_row=20))
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row if cell is not None]
        has_code = any('CODE' in c or 'INST' in c for c in normalized)
        has_name = any('SCHOOL' in c or 'NAME' in c or 'INSTITUTION' in c for c in normalized)
        if has_code and has_name:
            full_norm = [normalize_label(cell).upper() for cell in row]
            return idx, full_norm

        if has_code and not has_name and idx > 0:
            prev_row = rows[idx - 1]
            merged = merge_header_rows(row, prev_row)
            merged_norm = [val.upper() for val in merged]
            if any('SCHOOL' in c or 'NAME' in c or 'INSTITUTION' in c for c in merged_norm):
                return idx, merged

        if has_name and not has_code and idx + 1 < len(rows):
            next_row = rows[idx + 1]
            merged = merge_header_rows(next_row, row)
            merged_norm = [val.upper() for val in merged]
            if any('CODE' in c or 'INST' in c for c in merged_norm):
                return idx + 1, merged

    return None, None

def resolve_column_indices(header_row):
    """Resolve cell column positions for metadata fields from header row."""
    code_col, name_col, region_col, district_col, location_col, gender_col, type_col = None, None, None, None, None, None, None
    if not header_row:
        return code_col, name_col, region_col, district_col, location_col, gender_col, type_col

    for idx, col in enumerate(header_row):
        col_u = col.upper()
        if not col_u:
            continue
        if code_col is None and any(k in col_u for k in ['INST. CODE', 'SCHOOL CODE', 'INST CODE', 'CODE']) and 'POSTAL' not in col_u and 'DISTRICT' not in col_u:
            code_col = idx
        elif name_col is None and any(k in col_u for k in ['SCHOOL NAME', 'NAME OF SCHOOL', 'INSTITUTION', 'SCHOOL']) and 'TYPE' not in col_u and 'CODE' not in col_u:
            name_col = idx
        elif region_col is None and 'REGION' in col_u:
            region_col = idx
        elif district_col is None and any(k in col_u for k in ['DISTRICT', 'MUNICIPAL', 'METROPOLITAN', 'ASSEMBLY']):
            district_col = idx
        elif location_col is None and any(k in col_u for k in ['LOCATION', 'TOWN', 'CITY']):
            location_col = idx
        elif gender_col is None and any(k in col_u for k in ['GENDER', 'SEX', 'BOYS/GIRLS']):
            gender_col = idx
        elif type_col is None and any(k in col_u for k in ['TYPE OF SCHOOL', 'SCHOOL TYPE', 'CATEGORY OF SCH', 'TYPE']):
            type_col = idx

    return code_col, name_col, region_col, district_col, location_col, gender_col, type_col

def identify_single_program_label(col_candidates, col_idx=None):
    """Extract the most relevant single program label from a header-column candidate list."""
    cleaned_candidates = [normalize_label(c) for c in col_candidates if normalize_label(c)]
    if not cleaned_candidates:
        return None

    joined = ' '.join(cleaned_candidates).upper()

    if col_idx is not None and col_idx in range(7, 11) and any(token in joined for token in ['AGRIC', 'BUS', 'TECH', 'ECON', 'HOM']):
        grouped_labels = ['AGRICULTURAL SCIENCE', 'BUSINESS', 'TECHNICAL', 'HOME ECONOMICS']
        offset = col_idx - 7
        if 0 <= offset < len(grouped_labels):
            return grouped_labels[offset]

    if 'VIS' in joined or ('ARTS' in joined and 'GEN' not in joined):
        return 'VISUAL ARTS'
    if 'GEN' in joined and 'SCI' in joined:
        return 'GENERAL SCIENCE'
    if 'GEN' in joined or 'LANG' in joined:
        return 'GENERAL ARTS'
    if 'STEM' in joined:
        return 'STEM'
    if 'ECON' in joined or 'HOM' in joined:
        return 'HOME ECONOMICS'
    if 'TECH' in joined:
        return 'TECHNICAL'
    if 'AGRIC' in joined:
        return 'AGRICULTURAL SCIENCE'
    if 'BUS' in joined:
        return 'BUSINESS'

    # Fall back to the best explicit alias match.
    for candidate in cleaned_candidates:
        cand_u = candidate.upper().strip()
        if 'PROGRAMME' in cand_u and ('AGRIC' in cand_u or 'BUS' in cand_u or 'HOM' in cand_u or 'VIS' in cand_u or 'GEN' in cand_u):
            continue
        if cand_u.count('/') >= 3 or cand_u.count(',') >= 3 or len(cand_u.split()) >= 8:
            continue

        for std_name, aliases in STANDARD_PROGRAM_ALIASES:
            if any(alias == cand_u or f" {alias} " in f" {cand_u} " or cand_u.startswith(alias) for alias in aliases):
                return std_name

    for candidate in cleaned_candidates:
        cand_u = candidate.upper().strip()
        if cand_u in METADATA_IGNORE_HEADERS:
            continue
        if any(k in cand_u for k in ['PROGRAMME', 'SUBJECTS OFFERED', 'COURSES OFFERED']) and len(cand_u) > 25:
            continue
        if cand_u:
            return candidate.strip()
    return None

def parse_program_columns_multi_row(rows, header_index):
    """Inspect the multi-row header window to map marker columns to the correct program labels."""
    if header_index is None or not rows or header_index >= len(rows):
        return [], header_index + 1 if header_index is not None else 0

    max_cols = max(len(r) for r in rows)
    window_start = max(0, header_index - 1)
    window_end = min(len(rows), header_index + 3)

    program_columns = []
    for col_idx in range(max_cols):
        col_candidates = []
        for r_idx in range(window_start, window_end):
            row = rows[r_idx]
            if col_idx < len(row) and row[col_idx]:
                val = normalize_label(row[col_idx])
                if val and val.upper() not in METADATA_IGNORE_HEADERS:
                    col_candidates.append(val)

        if not col_candidates:
            continue

        prog_label = identify_single_program_label(col_candidates, col_idx)
        if prog_label and prog_label.upper() not in METADATA_IGNORE_HEADERS:
            program_columns.append((col_idx, prog_label))

    # If the header structure is compact (e.g. grouped labels like 'AGRIC BUS TECH HOM ECON'),
    # fall back to the specific code-row values below the header to recover the real program columns.
    if not program_columns:
        code_row = rows[header_index + 1] if header_index + 1 < len(rows) else []
        for col_idx in range(max_cols):
            if col_idx >= len(code_row):
                break
            if not code_row[col_idx]:
                continue
            if str(code_row[col_idx]).isdigit() and len(str(code_row[col_idx])) == 3:
                program_columns.append((col_idx, 'UNKNOWN'))

    # Determine real data start index (past sub-header rows)
    start_idx = header_index + 1
    while start_idx < len(rows):
        row = rows[start_idx]
        row_str = ' '.join(str(c) for c in row if c).upper()
        has_code = any(normalize_code(cell) is not None for cell in row)
        has_real_name = any(
            len(normalize_label(c)) > 3 and normalize_label(c).upper() not in METADATA_IGNORE_HEADERS
            and not any(s[0] in normalize_label(c).upper() for s in STANDARD_PROGRAM_ALIASES)
            for c in row if c
        )
        if has_code or has_real_name:
            break
        start_idx += 1

    return program_columns, start_idx

def parse_main_register_sheet(ws):
    """Parse school register worksheet using state tracking for region & district context."""
    header_index, header_row = find_header_row(ws)
    
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    if not rows:
        return []

    if header_index is not None and header_index < len(rows):
        h_row = header_row
        program_columns, start_idx = parse_program_columns_multi_row(rows, header_index)
        code_col, name_col, region_col, district_col, location_col, gender_col, type_col = resolve_column_indices(h_row)
    else:
        program_columns = []
        code_col, name_col, region_col, district_col, location_col, gender_col, type_col = None, None, None, None, None, None, None
        start_idx = 0

    sheet_category = None
    top_text = ' '.join(str(cell or '') for row in rows[:max(10, start_idx)] for cell in row if cell)
    for keyword, category in CATEGORY_KEYWORDS.items():
        if keyword in top_text.upper():
            sheet_category = category
            break

    sheet_region = infer_region_from_text(ws.title) or infer_region_from_text(top_text)
    current_region = sheet_region
    current_district = None

    schools = []
    for row in rows[start_idx:]:
        row_str = ' '.join(str(c) for c in row if c).strip()
        if not row_str:
            continue

        code = None
        found_code_col = None
        
        if code_col is not None and code_col < len(row):
            code = normalize_code(row[code_col])
            if code:
                found_code_col = code_col

        if not code:
            for idx, cell in enumerate(row):
                norm_c = normalize_code(cell)
                if norm_c:
                    code = norm_c
                    found_code_col = idx
                    break

        if not code:
            detected_region = infer_region_from_text(row_str)
            if detected_region and ('REGION' in row_str.upper() or len(row_str.split()) <= 4):
                current_region = detected_region
                current_district = None
                continue

            if any(k in row_str.upper() for k in ['DISTRICT:', 'MUNICIPAL:', 'METROPOLITAN:', 'DISTRICT ASSEMBLY', 'MUNICIPAL ASSEMBLY']):
                clean_dist = re.sub(r'(?i)^(DISTRICT|MUNICIPAL|METROPOLITAN|HEADER|SECTION)\s*[:\-]?\s*', '', row_str).strip()
                clean_dist = re.sub(r'(?i)\s*(DISTRICT|MUNICIPAL|METROPOLITAN)\s*$', '', clean_dist).strip()
                if clean_dist and len(clean_dist) < 60:
                    current_district = clean_dist
                continue
            continue

        name = None
        if name_col is not None and name_col < len(row):
            candidate = normalize_label(row[name_col])
            if candidate and candidate.upper() not in METADATA_IGNORE_HEADERS:
                if not re.match(r'^\d+$', candidate) and len(candidate) > 2:
                    name = candidate

        if not name:
            for idx, cell in enumerate(row):
                if idx == found_code_col:
                    continue
                cand = normalize_label(cell)
                if not cand or cand.upper() in METADATA_IGNORE_HEADERS:
                    continue
                if re.match(r'^\d+$', cand) or is_marker(cand):
                    continue
                if len(cand) >= 3 and not infer_type_from_value(cand) == 'TVET':
                    name = cand
                    break

        if not name or name.upper() in METADATA_IGNORE_HEADERS:
            continue

        row_region = normalize_label(row[region_col]) if region_col is not None and region_col < len(row) else ''
        if row_region:
            detected_row_reg = infer_region_from_text(row_region) or row_region
            if detected_row_reg and detected_row_reg != 'Unknown':
                current_region = detected_row_reg
                
        region = (row_region if row_region and row_region != 'Unknown' else None) or current_region or infer_region_from_code(code) or 'Unknown'

        row_district = normalize_label(row[district_col]) if district_col is not None and district_col < len(row) else ''
        if row_district and row_district.upper() not in {'UNKNOWN', 'DISTRICT', 'N/A', 'NONE'}:
            current_district = row_district

        district = row_district or current_district or (f"{location_col and row[location_col]} District" if location_col and location_col < len(row) and row[location_col] else f"{region} District")
        district = re.sub(r'(?i)^DISTRICT:\s*', '', str(district)).strip()
        if not district or district.upper() == 'UNKNOWN DISTRICT':
            district = f"{region} District" if region != 'Unknown' else 'Central District'

        location = normalize_label(row[location_col]) if location_col is not None and location_col < len(row) else ''
        gender = normalize_label(row[gender_col]) if gender_col is not None and gender_col < len(row) else ''
        row_type = normalize_label(row[type_col]) if type_col is not None and type_col < len(row) else ''
        
        school_type = infer_type_from_value(row_type) or infer_type_from_value(name) or infer_type_from_value(row[-1] if row else None) or 'SHS'

        if not location or location.upper() == 'UNKNOWN':
            location = f"{district}, {region}" if region != 'Unknown' and district != 'Unknown District' else (region if region != 'Unknown' else '')

        status = infer_status_from_row(row)
        program_names = []
        for col_idx, label in program_columns:
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            if label != 'UNKNOWN' and is_marker(value):
                program_names.append(label)
            elif label == 'UNKNOWN' and is_marker(value):
                # For compact grouped headers, infer the label from the nearest header text above.
                header_context = []
                for r_idx in range(max(0, header_index - 1), min(len(rows), header_index + 3)):
                    if col_idx < len(rows[r_idx]) and rows[r_idx][col_idx]:
                        header_context.append(normalize_label(rows[r_idx][col_idx]))
                inferred = identify_single_program_label(header_context, col_idx)
                if inferred:
                    program_names.append(inferred)

        schools.append({
            'code': code,
            'name': name,
            'region': region,
            'district': district,
            'location': location,
            'category': sheet_category or ('A' if school_type in {'TVET', 'STEM', 'SHTS'} else 'C'),
            'status': status,
            'gender': gender or 'Mixed',
            'type': school_type,
            'programNames': list(dict.fromkeys(program_names)),
            'progs': [],
        })
    return schools

def parse_appendix2_sheet(ws):
    """Parse Appendix 2 for program offerings mapping by school name."""
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'S/N' in normalized or 'SCHOOL' in normalized or 'ALL SENIOR HIGH' in str(normalized):
            header_index = idx
            break
    if header_index is None:
        return {}

    program_columns, start_idx = parse_program_columns_multi_row(rows, header_index)
    mapping = {}
    for row in rows[start_idx:]:
        if len(row) <= 2:
            continue
        name = normalize_label(row[2]) if len(row) > 2 else normalize_label(row[1])
        if not name or name.upper() in METADATA_IGNORE_HEADERS:
            continue
        programs = [
            label for idx, label in program_columns
            if idx < len(row) and is_marker(row[idx])
        ]
        if programs:
            key = normalize_school_name(name) or normalize_label(name).upper()
            mapping.setdefault(key, []).extend(programs)
            mapping[key] = list(dict.fromkeys(mapping[key]))
    return mapping

def parse_appendix4_sheet(ws):
    """Parse STEM specific worksheet data."""
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index, _ = find_header_row(ws)
    if header_index is None:
        return {}, []

    label_by_col = {
        7: 'BIO-MEDICAL SCIENCE',
        8: 'ENGINEERING SCIENCE',
        9: 'AVIATION & AEROSPACE ENGINEERING',
        10: 'COMPUTING',
        11: 'ROBOTICS',
        12: 'AGRICULTURAL SCIENCE',
        13: 'MANUFACTURING ENGINEERING',
    }
    mapping = {}
    sheet_schools = []
    current_code = None

    for row in rows[header_index + 1:]:
        code = normalize_code(row[3] if len(row) > 3 else '')
        if code:
            current_code = code
        if not current_code:
            continue

        name = normalize_label(row[4] if len(row) > 4 else '')
        reg_text = normalize_label(row[1] if len(row) > 1 else '')
        region = infer_region_from_text(reg_text) or infer_region_from_code(code) or 'Unknown'

        if code and name:
            school = {
                'code': code,
                'name': name,
                'region': region,
                'district': normalize_label(row[2] if len(row) > 2 else '') or f"{region} District",
                'location': normalize_label(row[5] if len(row) > 5 else '') or region,
                'category': 'A',
                'status': infer_status_from_row(row),
                'gender': normalize_label(row[6] if len(row) > 6 else '') or 'Mixed',
                'type': 'STEM',
                'programNames': [],
                'progs': [],
            }
            sheet_schools.append(school)

        programs = [
            label for col, label in label_by_col.items()
            if col < len(row) and is_marker(row[col])
        ]
        if programs:
            mapping.setdefault(current_code, []).extend(programs)
            mapping[current_code] = list(dict.fromkeys(mapping[current_code]))

    return mapping, sheet_schools

def parse_appendix1_sheet(ws):
    """Parse TVET specific worksheet data."""
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'S/N' in normalized and any(k in normalized for k in ['INST. CODE', 'CODE', 'INST CODE']):
            header_index = idx
            break
    if header_index is None:
        return {}

    header = [normalize_label(cell).upper() for cell in rows[header_index]]
    code_col = next((i for i, h in enumerate(header) if any(k in h for k in ['INST. CODE', 'CODE', 'INST CODE'])), 1)
    
    mapping = {}
    for row in rows[header_index + 1:]:
        code = normalize_code(row[code_col] if code_col < len(row) else '')
        if not code:
            continue
        for idx, cell in enumerate(row):
            if idx <= code_col:
                continue
            if is_marker(cell):
                label = normalize_label(rows[header_index][idx] if idx < len(rows[header_index]) else '')
                if label and not re.fullmatch(r'\d{3,4}', label):
                    mapping.setdefault(code, []).append(label)
        if code in mapping:
            mapping[code] = list(dict.fromkeys(mapping[code]))
    return mapping

def merge_program_map(*program_lists):
    """Merge multiple program-label lists while preserving the full set of distinct programs."""
    merged = []
    for program_list in program_lists:
        for name in program_list or []:
            if not name:
                continue
            normalized = normalize_label(name).upper()
            if normalized in {normalize_label(item).upper() for item in merged}:
                continue
            merged.append(normalize_label(name))
    return merged


def canonical_progs_for_program_names(names):
    """Map explicit program names to standardized GES shortcodes."""
    canonical = []
    for value in names:
        normalized = normalize_label(value).upper()
        if not normalized:
            continue

        # Match explicit program names first
        matched = False
        for std_name, aliases in STANDARD_PROGRAM_ALIASES:
            if any(alias == normalized or f" {alias} " in f" {normalized} " or normalized.startswith(alias) for alias in aliases):
                if std_name == 'AGRICULTURAL SCIENCE':
                    canonical.append('AGRIC')
                elif std_name == 'BUSINESS':
                    canonical.append('BUS')
                elif std_name == 'HOME ECONOMICS':
                    canonical.append('HOM. ECON.')
                elif std_name == 'VISUAL ARTS':
                    canonical.append('VIS. ARTS')
                elif std_name == 'GENERAL ARTS':
                    canonical.append('GEN. ARTS')
                elif std_name == 'GENERAL SCIENCE':
                    canonical.append('GEN. SCI')
                elif std_name == 'TECHNICAL':
                    canonical.append('TECH')
                elif std_name == 'STEM':
                    canonical.append('STEM')
                matched = True
                break

        if matched:
            continue

        # Regex fallback for non-standard or compound TVET/STEM names
        if re.search(r'\b(STEM|BIO|ENGINEERING|COMPUTING|ROBOTICS|AEROSPACE)\b', normalized):
            canonical.append('STEM')
        elif re.search(r'\b(TECH|TVET|VOCATIONAL|TRADE|TRADES|CONSTRUCTION|AUTOMOTIVE|MECHANICAL|ELECTRICAL|HOSPITALITY|COSMETOLOGY)\b', normalized):
            canonical.append('TECH')
        elif re.search(r'\b(AGRIC|AGRO|CROP|ANIMAL)\b', normalized):
            canonical.append('AGRIC')
        elif re.search(r'\b(BUS|ACCOUNTING|FINANCE|MARKETING)\b', normalized):
            canonical.append('BUS')
        elif re.search(r'\b(HOME\s*ECON|HOM\. ECON|FOOD|TEXTILES|CATERING)\b', normalized):
            canonical.append('HOM. ECON.')
        elif re.search(r'\b(VIS|VISUAL)\s*ARTS?\b', normalized):
            canonical.append('VIS. ARTS')
        elif re.search(r'\b(GEN(?:ERAL)?\s*ARTS?)\b', normalized):
            canonical.append('GEN. ARTS')
        elif re.search(r'\b(GEN(?:ERAL)?\s*SCI(?:ENCE)?)\b', normalized):
            canonical.append('GEN. SCI')

    return list(dict.fromkeys(canonical))

def main():
    print(f"Loading workbook from {xlsx_path}...")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)

    school_by_code = {}
    
    # Iterate across all worksheets without dropping sheets
    for ws in wb.worksheets:
        schools = parse_main_register_sheet(ws)
        for school in schools:
            code = school['code']
            if code in school_by_code:
                existing = school_by_code[code]
                existing['programNames'] = merge_program_map(existing.get('programNames', []), school.get('programNames', []))
                if existing['region'] == 'Unknown' and school['region'] != 'Unknown':
                    existing['region'] = school['region']
                if 'District' in existing['district'] and existing['district'] != school['district'] and school['district'] != 'Unknown District':
                    existing['district'] = school['district']
            else:
                school_by_code[code] = school

    appendix2_by_name = {}
    appendix4_by_code = {}
    appendix4_extra_schools = []
    appendix1_by_code = {}

    for ws in wb.worksheets:
        header_text = ' '.join(str(cell or '') for row in ws.iter_rows(values_only=True, max_row=8) for cell in row).upper()
        if 'APPENDIX 2' in header_text:
            appendix2_by_name.update(parse_appendix2_sheet(ws))
        if 'APPENDIX 4' in header_text or 'STEM SCHOOLS' in header_text:
            codes, extra = parse_appendix4_sheet(ws)
            appendix4_by_code.update(codes)
            appendix4_extra_schools.extend(extra)
        if 'APPENDIX 1' in header_text:
            appendix1_by_code.update(parse_appendix1_sheet(ws))

    # Merge additional programs from appendices.
    # Prefer exact code-based matches first, then exact normalized name matches.
    for school in school_by_code.values():
        code = school['code']
        name_key = normalize_school_name(school['name'])
        additional_programs = []
        if code in appendix1_by_code and school['type'] == 'TVET':
            additional_programs.extend(appendix1_by_code[code])
        if code in appendix4_by_code:
            additional_programs.extend(appendix4_by_code[code])
        if name_key and name_key in appendix2_by_name:
            additional_programs.extend(appendix2_by_name[name_key])
        school['programNames'] = merge_program_map(school.get('programNames', []), additional_programs)

    # Merge STEM/TVET extra schools if omitted from main register
    for extra_school in appendix4_extra_schools:
        code = extra_school['code']
        if code not in school_by_code:
            school_by_code[code] = extra_school
        else:
            school_by_code[code]['programNames'] = merge_program_map(
                school_by_code[code].get('programNames', []),
                appendix4_by_code.get(code, []),
            )

    # Finalize region & district fallbacks using 3-digit school codes
    for school in school_by_code.values():
        if school['region'] == 'Unknown':
            code_region = infer_region_from_code(school['code'])
            if code_region:
                school['region'] = code_region

        if not school.get('programNames'):
            if school['type'] == 'TVET':
                school['programNames'] = ['TECHNICAL PROGRAMMES']
            elif school['type'] == 'STEM':
                school['programNames'] = ['STEM']
            else:
                school['programNames'] = ['GENERAL SCIENCE']

        canonical = canonical_progs_for_program_names(school['programNames'])
        if school['type'] == 'TVET':
            school['progs'] = ['TECH']
        elif school['type'] == 'STEM':
            school['progs'] = ['STEM'] + [p for p in canonical if p != 'STEM']
        else:
            school['progs'] = canonical or ['GEN. SCI']

    # Sort all schools by school code
    all_schools = sorted(school_by_code.values(), key=lambda s: (s.get('code') or '', s.get('name') or ''))

    for school in all_schools:
        school['location'] = normalize_label(school['location'])
        if school['region'] == 'Unknown':
            code_region = infer_region_from_code(school['code'])
            if code_region:
                school['region'] = code_region
        school['district'] = normalize_label(school['district'])
        if not school['district'] or school['district'] == 'Unknown District':
            school['district'] = f"{school['region']} District" if school['region'] != 'Unknown' else 'Central District'

    schools_by_region = defaultdict(list)
    for school in all_schools:
        schools_by_region[school['region']].append(school)

    # Write output JSON files
    with (output_dir / 'schools_all.json').open('w', encoding='utf-8') as f:
        json.dump(all_schools, f, ensure_ascii=False, indent=2)

    with (output_dir / 'schools_by_region.json').open('w', encoding='utf-8') as f:
        json.dump({region: schools for region, schools in sorted(schools_by_region.items())}, f, ensure_ascii=False, indent=2)

    summary = {
        'total_schools': len(all_schools),
        'regions': {region: len(schools) for region, schools in sorted(schools_by_region.items())},
        'school_types': {
            label: sum(1 for s in all_schools if s['type'] == label)
            for label in sorted({s['type'] for s in all_schools})
        },
    }
    with (output_dir / 'schools_summary.json').open('w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f'Successfully wrote {len(all_schools)} schools to {output_dir / "schools_all.json"}')

if __name__ == '__main__':
    main()