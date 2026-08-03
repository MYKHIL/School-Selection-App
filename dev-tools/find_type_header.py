import openpyxl, re
from pathlib import Path
path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('sheetcount', len(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    print('\nSheet:', name)
    found = False
    for r in range(1, 25):
        row = [ws.cell(row=r, col=c).value for c in range(1, 40)]
        if any(cell is not None and re.search(r'\bTYPE\b|\bSCH\. TYPE\b|\bSCHOOL TYPE\b|\bSCH TYPE\b|\bSHTS\b|\bSTEM\b|\bTVET\b', str(cell).upper()) for cell in row):
            print('row', r, [str(cell).strip() if cell is not None else '' for cell in row])
            found = True
    if not found:
        print('  no explicit type/search header rows found')
