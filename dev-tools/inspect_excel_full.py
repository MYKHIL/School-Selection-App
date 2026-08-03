import openpyxl
from pathlib import Path

path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('sheetcount', len(wb.sheetnames))
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    print('\n=== Sheet', idx, repr(name), '===')
    nonempty = False
    for r in range(1, 61):
        row = [ws.cell(row=r, column=c).value for c in range(1, 61)]
        if any(cell is not None for cell in row):
            nonempty = True
            print(r, [str(cell).strip() if cell is not None else '' for cell in row])
            if r >= 25:
                break
    if not nonempty:
        print('  empty first 60 rows')
