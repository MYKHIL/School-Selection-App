from openpyxl import load_workbook
import re

wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)

sheet = wb['Sheet37']
rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in sheet.iter_rows(values_only=True)]

for code in ['9050101','9060200','9050301','9090102']:
    print('CODE', code)
    # parse sheet as JS logic
    def normalize_cell(value):
        return (value or '').strip()
    def is_header_row(row):
        values = [normalize_cell(v) for v in row if normalize_cell(v)]
        numeric_count = sum(1 for v in values if re.match(r'^\d{3,4}$', v))
        return numeric_count >= 5
    header_idx = None
    for idx,row in enumerate(rows):
        if is_header_row(row):
            header_idx = idx
            break
    if header_idx is None:
        print('no header')
        continue
    label_row = rows[header_idx-1]
    code_row = [normalize_cell(v) for v in rows[header_idx]]
    program_cols = [ci for ci,v in enumerate(code_row) if re.match(r'^\d{3,4}$', v)]
    program_header_map = {}
    ignore_pattern = re.compile(r'^(?:S/N|REGION|DISTRICT|INST\.? CODE|INSTITUTION|LOCATION|GENDER|STATUS|CATEGORY|NO\.? OF|APPENDIX)$', re.I)
    for ci in program_cols:
        header = normalize_cell(label_row[ci]) or code_row[ci]
        if header and not ignore_pattern.match(header):
            program_header_map[ci] = header
    print('header count', len(program_header_map))
    programs=[]
    for row in rows:
        row_data = [normalize_cell(v) for v in row]
        if code not in row_data:
            continue
        code_index = row_data.index(code)
        for ci,cell_value in enumerate(row_data):
            if ci == code_index: continue
            if not cell_value: continue
            if ci in program_header_map and re.match(r'^(?:X|x|?|?|1|Y|YES)$', cell_value):
                programs.append(program_header_map[ci])
                continue
            if re.match(r'^appendix', cell_value, re.I):
                continue
            if re.search(r'refer to|detailed|programme?s?', cell_value, re.I):
                continue
            if re.match(r'^(?:MIXED|MALES|FEMALES|DAY/BOARDING|DAY|BOARDING|TVET|SHTS|SHS)$', cell_value, re.I):
                continue
            if re.match(r'^\d{1,4}$', cell_value):
                continue
            if re.match(r'^(?:[A-Z]{1,3}\.|\w+\s+INSTITUTE?)$', cell_value, re.I) and ci < code_index:
                continue
            programs.append(cell_value)
        if programs:
            break
    print('programs', programs)
