import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
xlsx = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'


def normalize_space(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_label(value):
    value = normalize_space(value)
    value = value.replace(' - ', '-').replace('  ', ' ')
    return value.strip(' ,;:-')


def is_marker(value):
    marker = normalize_space(value).upper()
    return marker in {'X', 'XX', '1', 'Y', 'YES', '✔', '✓'}


def parse_shs_shts_appendix(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'S/N' in normalized and 'ALL SENIOR HIGH/ TECH SCHOOLS' in normalized:
            header_index = idx
            break
    return header_index, rows


wb = load_workbook(xlsx, data_only=True, read_only=True)
for idx, sheet in enumerate(wb.worksheets, start=1):
    if any(isinstance(cell, str) and 'APPENDIX 2' in cell.upper() for row in sheet.iter_rows(values_only=True, max_row=8) for cell in row):
        print('===', idx, sheet.title)
        header_index, rows = parse_shs_shts_appendix(sheet)
        print('header_index', header_index)
        if header_index is not None:
            print('header row', rows[header_index])
            for i in range(header_index+1, header_index+6):
                print(i+1, rows[i])
        else:
            for i,row in enumerate(rows[:15], start=1):
                print(i, row)
        print('---')

# Search for target names in all Appendix 2 sheets
searches = ['MANCEL GIRLS', 'SUNYANI SENIOR HIGH', 'ACCRA SENIOR HIGH', 'KNUST SENIOR HIGH']
for term in searches:
    print('SEARCH TERM', term)
    for idx, sheet in enumerate(wb.worksheets, start=1):
        if any(isinstance(cell, str) and 'APPENDIX 2' in cell.upper() for row in sheet.iter_rows(values_only=True, max_row=8) for cell in row):
            found = []
            for i,row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if any(isinstance(cell, str) and term in cell.upper() for cell in row if cell):
                    found.append((i, row))
            if found:
                print('  sheet', idx, sheet.title, 'matches', len(found))
                for pos, row in found[:3]:
                    print('   ', pos, row)
    print()