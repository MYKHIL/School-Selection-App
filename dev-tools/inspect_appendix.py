from openpyxl import load_workbook
from pathlib import Path
import re
p = Path(r'd:/School Selection App/scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = load_workbook(p, data_only=True)
print('sheets', len(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True))
    text = '\n'.join(' '.join(str(cell or '') for cell in row) for row in rows)
    if re.search(r'APPENDIX|PROGRAMME|PROGRAM', text, re.I):
        print('----', name, 'rows', ws.max_row, 'cols', ws.max_column)
        for r,row in enumerate(rows,1):
            print(r, row)
        print()
