import json
import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
json_path = root / 'data' / 'schools_all.json'
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'


def normalize_space(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_label(value):
    value = normalize_space(value)
    value = value.replace(' - ', '-').replace('  ', ' ')
    return value.strip(' ,;:-')


def normalize_school_name(value):
    text = normalize_label(value or '')
    text = re.sub(r"\b(SENIOR HIGH/TECH|SENIOR HIGH TECH|SENIOR HIGH|SR\. HIGH|SNR HIGH|SNR\. HIGH|TECH SCHOOL|TECHNICAL SCHOOL|TECH|SHTS|SHS|STEM|TVET|ACADEMY|INSTITUTE|SCHOOL)\b", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'[^A-Z0-9 ]+', ' ', text.upper())
    return re.sub(r'\s+', ' ', text).strip()


def sheet_contains_appendix(ws, appendix_text):
    for row in ws.iter_rows(values_only=True, max_row=8):
        for cell in row:
            if cell and appendix_text in str(cell).upper():
                return True
    return False


def parse_shs_shts_appendix(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'S/N' in normalized and 'ALL SENIOR HIGH/ TECH SCHOOLS' in normalized:
            header_index = idx
            break
    if header_index is None:
        return []

    header_row = rows[header_index]
    subject_headers = [(col_idx, normalize_label(header_row[col_idx])) for col_idx in range(5, len(header_row)) if normalize_label(header_row[col_idx])]
    mapping = {}
    for row in rows[header_index + 1:]:
        if not row:
            continue
        name = normalize_label(row[2]) if len(row) > 2 else ''
        if not name:
            continue
        programs = [label for col_idx, label in subject_headers if col_idx < len(row) and str(row[col_idx]).strip().upper() in {'X','XX','1','Y','YES','✔','✓'}]
        if programs:
            key = normalize_school_name(name) or normalize_label(name).upper()
            mapping[key] = list(dict.fromkeys(mapping.get(key, []) + programs))
    return mapping


def all_appendix2_rows(ws):
    return [' '.join(str(cell or '').upper() for cell in row) for row in ws.iter_rows(values_only=True)]


with json_path.open('r', encoding='utf-8') as f:
    schools = json.load(f)

wb = load_workbook(xlsx_path, data_only=True, read_only=True)
appendix2_sheets = [ws for ws in wb.worksheets if sheet_contains_appendix(ws, 'APPENDIX 2')]
print('Appendix2 sheets', [ws.title for ws in appendix2_sheets])
appendix2_rows = []
for ws in appendix2_sheets:
    appendix2_rows.extend(all_appendix2_rows(ws))

incomplete = [s for s in schools if s.get('type') in ('SHS', 'SHTS') and len([x for x in s.get('programNames', []) if normalize_label(x)]) <= 1]
print('incomplete SHS/SHTS count', len(incomplete))

sample = []
exact_match = 0
partial_match = 0
for s in incomplete[:200]:
    name = s.get('name', '')
    key = normalize_school_name(name)
    if key and any(key in row for row in appendix2_rows):
        exact_match += 1
        continue
    tokens = set(key.split())
    found = []
    for row in appendix2_rows:
        row_norm = re.sub(r'[^A-Z0-9 ]+', ' ', row)
        row_norm = re.sub(r'\s+', ' ', row_norm).strip()
        row_tokens = set(row_norm.split())
        common = tokens & row_tokens
        if len(common) >= min(3, len(tokens)) and key not in row_norm:
            found.append((len(common), row[:120]))
    if found:
        partial_match += 1
    sample.append((name, key, len(found), found[:3]))

print('exact normalized matches within first 200 incomplete:', exact_match)
print('partial candidate matches within first 200 incomplete:', partial_match)
print('unmatched sample count:', sum(1 for _,_,count,_ in sample if count == 0))
for name,key,count,found in sample[:30]:
    if count == 0:
        print('NO MATCH', name, key)
print('DONE')
