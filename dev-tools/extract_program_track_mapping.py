#!/usr/bin/env python3
"""
Extract programme canonical -> track mappings by scanning the register workbook
and matching known programme aliases found in data/programmes.json against
Appendix sheets (Appendix 1 -> TVET, Appendix 2 -> SHTS, Appendix 4 -> STEM).
Writes `data/programme_track_map.json`.
"""
import json
import re
from openpyxl import load_workbook

BASE_XLSX = 'scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
PROG_JSON = 'data/programmes.json'
OUT_JSON = 'data/programme_track_map.json'

def load_program_definitions():
    with open(PROG_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)

def build_alias_map(defs):
    entries = []
    for item in defs.get('programmes', []):
        canonical = item.get('canonical')
        display = item.get('display')
        aliases = list(item.get('aliases', []))
        if display:
            aliases.append(display)
        aliases = [str(a).strip() for a in aliases if a]
        entries.append((canonical, set(a.upper() for a in aliases)))
    for item in defs.get('expandedPrograms', []):
        canonical = item.get('canonical')
        display = item.get('display')
        aliases = list(item.get('aliases', []))
        if display:
            aliases.append(display)
        aliases = [str(a).strip() for a in aliases if a]
        # include expanded entries as separate display aliases
        entries.append((canonical, set(a.upper() for a in aliases)))

    # flatten to map alias -> canonical
    alias_map = {}
    for canonical, aliasset in entries:
        for alias in aliasset:
            alias_map[alias] = canonical
    return alias_map

def classify_sheet(name, rows_text):
    text = '\n'.join(rows_text).upper()
    if 'APPENDIX 1' in text:
        return 'TVET'
    if 'APPENDIX 2' in text or 'SENIOR HIGH TECH' in text or 'SHTS' in text or 'TECHNICAL' in text:
        return 'SHTS'
    if 'APPENDIX 4' in text or 'STEM' in text:
        return 'STEM'
    return None

def scan_workbook():
    wb = load_workbook(BASE_XLSX, data_only=True)
    sheet_map = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows_text = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            row_txt = ' '.join(str(c).strip() for c in row if c is not None)
            if row_txt.strip():
                rows_text.append(row_txt)
        kind = classify_sheet(name, rows_text)
        if kind:
            sheet_map[name] = (kind, rows_text)
    return wb, sheet_map

def map_aliases_to_tracks(wb, sheet_map, alias_map):
    mapping = {}
    # initialize mapping for all canonicals
    for canon in set(alias_map.values()):
        mapping[canon] = set()

    # search only sheets that were classified as relevant
    for sheetname, (track, rows_text) in sheet_map.items():
        ws = wb[sheetname]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if not cell:
                    continue
                txt = str(cell).strip().upper()
                # check tokenized alias matches
                for alias, canon in alias_map.items():
                    if alias in txt:
                        mapping.setdefault(canon, set()).add(track)
    return mapping

def save_mapping(mapdict):
    out = {k: sorted(list(v)) for k, v in mapdict.items() if v}
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump({'generated': True, 'mappings': out}, f, indent=2, ensure_ascii=False)
    print('Wrote', OUT_JSON)

def main():
    defs = load_program_definitions()
    alias_map = build_alias_map(defs)
    wb, sheet_map = scan_workbook()
    mapping = map_aliases_to_tracks(wb, sheet_map, alias_map)
    save_mapping(mapping)

if __name__ == '__main__':
    main()
