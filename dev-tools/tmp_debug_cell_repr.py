from openpyxl import load_workbook
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)
sheet = wb['Sheet37']
rows = list(sheet.iter_rows(values_only=True))
for idx,row in enumerate(rows, start=1):
    if row and any(str(cell).strip() == '9050101' for cell in row if cell is not None):
        print('ROW', idx)
        print([repr(cell) for cell in row[:30]])
        break

for idx,row in enumerate(rows, start=1):
    if row and any(str(cell).strip() == '9050301' for cell in row if cell is not None):
        print('ROW 9050301', idx)
        print([repr(cell) for cell in row[:30]])
        break

# print row 7 and 6 for header
for idx in [6,7,42]:
    row = rows[idx-1]
    print('INDEX', idx, [repr(cell) for cell in row[:30]])
