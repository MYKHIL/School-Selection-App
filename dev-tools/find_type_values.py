import openpyxl, re
from pathlib import Path
path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('sheetcount', len(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    print('\nSheet:', name)
    found = False
    for r in range(1, 40):
        row = [ws.cell(row=r, column=c).value for c in range(1, 50)]
        if any(cell is not None and isinstance(cell, str) and re.search(r'\b(SHTS|STEM|TVET|SHS|TYPE|SCHOOL TYPE|SCH\. TYPE)\b', cell, re.I) for cell in row):
            print('row', r, [str(cell).strip() if cell is not None else '' for cell in row])
            found = True
            if r > 20:
                break
    if not found:
        print('  no explicit type/search terms found in first 40 rows')
