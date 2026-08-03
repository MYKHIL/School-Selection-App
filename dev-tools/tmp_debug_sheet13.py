from openpyxl import load_workbook
from pathlib import Path
from scripts.export_school_register_json import find_header_row, parse_program_columns_multi_row, resolve_column_indices, normalize_label, normalize_code, infer_region_from_text, infer_region_from_code, parse_main_register_sheet

wb = load_workbook(Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'), data_only=True, read_only=True)
ws = wb['Sheet13']
rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
for idx, row in enumerate(rows[:20], start=1):
    print(idx, row)
print('---')
header_index, header_row = find_header_row(ws)
print('header_index', header_index)
print('header_row', header_row)
program_columns, start_idx = parse_program_columns_multi_row(rows, header_index)
print('program_columns', program_columns)
print('start_idx', start_idx)

row = rows[16]
print('row17 raw', row)
print('region cell', row[1], 'name cell', row[4], 'code cell', row[3])
print('normalize_code', normalize_code(row[3]))
print('infer_region_from_text', infer_region_from_text(row[1]))
print('infer_region_from_code', infer_region_from_code(normalize_code(row[3])))

parsed = parse_main_register_sheet(ws)
print('found count', len(parsed))
for s in parsed:
    if isinstance(s['name'], str) and 'accra stem academy' in s['name'].lower():
        print('parsed', s)
