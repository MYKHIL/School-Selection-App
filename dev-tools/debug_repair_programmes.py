import json
import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
json_path = root / 'data' / 'schools_all.json'


def normalize_space(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_label(value):
    value = normalize_space(value)
    value = value.replace(' - ', '-').replace('  ', ' ')
    return value.strip(' ,;:-')


def normalize_school_name(value):
    text = normalize_label(value or '')
    text = re.sub(r"\b(SENIOR HIGH/TECH|SENIOR HIGH TECH|SENIOR HIGH|SR\. HIGH|SNR HIGH|TECH SCHOOL|TECH|SHTS|STEM|TVET|ACADEMY|INSTITUTE|SCHOOL)\b", ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'[^A-Z0-9 ]+', ' ', text.upper())
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def is_marker(value):
    marker = normalize_space(value).upper()
    return marker in {'X', 'XX', '1', 'Y', 'YES', '✔', '✓'}


def parse_shs_shts_appendix(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_label(cell).upper() for cell in row]
        if 'S/N' in normalized and 'ALL SENIOR HIGH/ TECH SCHOOLS' in normalized:
            header_index = idx
            break
    if header_index is None:
        return {}

    header_row = rows[header_index]
    subject_headers = []
    for col_idx in range(5, len(header_row)):
        label = normalize_label(header_row[col_idx]) if col_idx < len(header_row) else ''
        if label:
            subject_headers.append((col_idx, label))
    if not subject_headers:
        return {}

    mapping = {}
    for row in rows[header_index + 1:]:
        if not row:
            continue
        name = normalize_label(row[2]) if len(row) > 2 else ''
        if not name:
            continue
        programs = []
        for col_idx, label in subject_headers:
            if col_idx < len(row) and is_marker(row[col_idx]):
                programs.append(label)
        if programs:
            mapping[normalize_label(name).upper()] = programs
    return mapping


def parse_tvet_sheet(ws):
    rows = [[normalize_space(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_index = None
    for idx, row in enumerate(rows):
        if any(cell.upper() in {'S/N', 'SN'} for cell in row[:6] if cell):
            header_index = idx
            break
    if header_index is None:
        return {}

    header_row = rows[header_index]
    code_row = rows[header_index + 1] if header_index + 1 < len(rows) else []
    mapping = {}
    for row in rows[header_index + 2:]:
        if not row:
            continue
        code = normalize_space(row[3]) if len(row) > 3 else ''
        if not re.fullmatch(r'\d{7}', code):
            continue
        for col_idx in range(7, len(row)):
            if not is_marker(row[col_idx]):
                continue
            label = normalize_label(header_row[col_idx]) if col_idx < len(header_row) else ''
            code_value = normalize_space(code_row[col_idx]) if col_idx < len(code_row) else ''
            if label and not re.fullmatch(r'\d{3,4}', label):
                mapping.setdefault(code, []).append(f"{label} ({code_value})" if re.fullmatch(r'\d{3,4}', code_value) else label)
    return mapping


def merge_programme_maps(target, source):
    for key, values in source.items():
        target.setdefault(key, []).extend(values)
        target[key] = list(dict.fromkeys(target[key]))


def sheet_contains_appendix(ws, appendix_text):
    for row in ws.iter_rows(values_only=True, max_row=8):
        for cell in row:
            if cell and appendix_text in str(cell).upper():
                return True
    return False


wb = load_workbook(xlsx_path, data_only=True, read_only=True)
code_mapping = {}
name_mapping = {}
for sheet in wb.worksheets:
    title = sheet.title
    if sheet_contains_appendix(sheet, 'APPENDIX 1'):
        merge_programme_maps(code_mapping, parse_tvet_sheet(sheet))
    if sheet_contains_appendix(sheet, 'APPENDIX 2'):
        merge_programme_maps(name_mapping, parse_shs_shts_appendix(sheet))
    if sheet_contains_appendix(sheet, 'APPENDIX 4') or 'STEM' in title.upper():
        merge_programme_maps(code_mapping, parse_tvet_sheet(sheet))

print('Appendix1 code keys', len(code_mapping))
print('Appendix2 name keys', len(name_mapping))
print('Example name keys', list(name_mapping.keys())[:20])

with json_path.open('r', encoding='utf-8') as f:
    schools = json.load(f)

found = 0
missing = 0
for code, name in [('0050109','KNUST Senior High'), ('0050159', "Mancell Girls' Senior High/Tech. Oduom"), ('0021315','Abomosu STEM Senior High')]:
    print('CODE', code, 'JSON', next((s for s in schools if s.get('code')==code), None))
for query in ['ACCRA SENIOR HIGH', 'KNUST SENIOR HIGH', 'MANCEL GIRLS', 'ABOMOSU STEM', 'SUNYANI SENIOR HIGH']:
    normalized = normalize_label(query).upper()
    if normalized in name_mapping:
        print('NAME MATCH', query, '->', name_mapping[normalized])
    else:
        matches = [k for k in name_mapping if normalized in k or k in normalized]
        print('NAME SEARCH', query, 'found', len(matches), 'matches', matches[:10])

print('SAMPLE names from mapping:')
for k in list(name_mapping.keys())[:50]:
    print(k)
