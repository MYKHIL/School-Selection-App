from openpyxl import load_workbook
import re
import os
path = os.path.join('scripts', 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
print('PATH', path, os.path.exists(path))
wb = load_workbook(path, data_only=True)
print('SHEETS', len(wb.sheetnames))
for idx, name in enumerate(wb.sheetnames, start=1):
    print(idx, repr(name))
print('--- scanning for appendix-like content ---')
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else '' for c in row])
    sample = ' '.join(' '.join(row).upper() for row in rows[:8])
    if 'APPENDIX' in sample or 'PROGRAMME' in sample or 'DETAILED' in sample or 'TECHNICAL' in sample or 'INDUSTRIAL' in sample:
        print('APPENDIX-LIKE SHEET', name)
        for i, row in enumerate(rows[:15], start=1):
            print(i, row)
        for code in ['9050101','9050301','9060200']:
            for i,row in enumerate(rows, start=1):
                if code in row:
                    print('FOUND CODE', code, 'at row', i)
                    break
        print('---')
