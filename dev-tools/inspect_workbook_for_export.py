from openpyxl import load_workbook
from pathlib import Path

path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
print('sheet_count', len(wb.sheetnames))
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    nonempty = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        rowvals = [str(v) if v is not None else '' for v in row[:12]]
        if any(v.strip() for v in rowvals):
            nonempty.append(rowvals)
    if nonempty:
        print('--- SHEET', idx, name, 'rows', ws.max_row, 'cols', ws.max_column)
        for row in nonempty[:8]:
            print(row)
