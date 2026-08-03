import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'


def normalize_space(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


wb = load_workbook(xlsx_path, data_only=True, read_only=True)
for idx, ws in enumerate(wb.worksheets, start=1):
    first8 = '\n'.join(' '.join(str(c or '') for c in row).upper() for row in ws.iter_rows(values_only=True, max_row=8))
    if 'APPENDIX 4' in first8 or 'STEM' in ws.title.upper() or any('APPENDIX 4' in str(c).upper() for row in ws.iter_rows(values_only=True, max_row=20) for c in row if c):
        print('=== sheet', idx, ws.title)
        for row_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=20), start=1):
            print(row_idx, [normalize_space(c) for c in row])
        print()
