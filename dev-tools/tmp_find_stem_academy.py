from openpyxl import load_workbook
import re
from pathlib import Path
wb = load_workbook(Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'), data_only=True, read_only=True)
needle = re.compile(r'accra stem academy', re.I)
for ws in wb.worksheets:
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(isinstance(cell, str) and needle.search(cell) for cell in row):
            print('FOUND in', ws.title, 'row', idx)
            print(row)
            # print header rows around it
            start = max(1, idx-4)
            for j in range(start, idx+2):
                print(j, list(ws.iter_rows(min_row=j, max_row=j, values_only=True))[0])
            raise SystemExit
print('NOT FOUND')
