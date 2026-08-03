from openpyxl import load_workbook
from pathlib import Path
from scripts.export_school_register_json import parse_appendix4_sheet, parse_appendix1_sheet, parse_appendix2_sheet, normalize_school_name
wb = load_workbook(Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'), data_only=True, read_only=True)
for ws in wb.worksheets:
    header_text = ' '.join(str(cell or '') for row in ws.iter_rows(values_only=True, max_row=8) for cell in row).upper()
    if 'APPENDIX 4' in header_text or 'STEM SCHOOLS' in header_text:
        codes, extra = parse_appendix4_sheet(ws)
        print('Sheet', ws.title, 'codes', len(codes), 'extra', len(extra))
        if '0012305' in codes:
            print('FOUND 0012305 in Appendix4 codes', codes['0012305'])
        for code, names in codes.items():
            if code.endswith('305'):
                print('maybe', code, names)
for ws in wb.worksheets:
    if any('APPENDIX 1' in str(cell).upper() for row in ws.iter_rows(values_only=True, max_row=8) for cell in row if cell):
        m = parse_appendix1_sheet(ws)
        if '0012305' in m:
            print('appendix1 0012305', m['0012305'])
for ws in wb.worksheets:
    if any('APPENDIX 2' in str(cell).upper() for row in ws.iter_rows(values_only=True, max_row=8) for cell in row if cell):
        m = parse_appendix2_sheet(ws)
        for key, names in list(m.items())[:20]:
            if 'ACC' in key or 'STEM' in key:
                print('Appendix2 key', key, names)
