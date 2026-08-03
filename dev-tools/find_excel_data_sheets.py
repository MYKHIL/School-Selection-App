import openpyxl, re
from pathlib import Path

path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('sheetcount', len(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    header_rows = []
    for r in range(1, 30):
        row = [ws.cell(row=r, column=c).value for c in range(1, 40)]
        textrow = [str(cell).strip().upper() if cell is not None else '' for cell in row]
        if 'CODE' in textrow and 'SCHOOL NAME' in textrow and 'LOCATION' in textrow:
            header_rows.append((r, textrow))
    if header_rows:
        print('\nSheet:', name)
        for r, textrow in header_rows:
            print(' header row', r, textrow)
        # print subsequent data rows
        start = header_rows[0][0]
        for r in range(start, start+15):
            row = [ws.cell(row=r, column=c).value for c in range(1, 45)]
            if any(cell is not None for cell in row):
                print(r, [str(cell).strip() if cell is not None else '' for cell in row])
