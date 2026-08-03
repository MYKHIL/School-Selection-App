import openpyxl
from pathlib import Path
from itertools import islice

path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
print('exists', path.exists())
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('sheets', wb.sheetnames)
for name in wb.sheetnames:
    sheet = wb[name]
    rows = list(sheet.iter_rows(values_only=True))
    nonempty = [row for row in rows if any(cell is not None for cell in row)]
    print('sheet', name, 'rows', len(rows), 'nonempty', len(nonempty))
    print(nonempty[:5])
    print('---')
