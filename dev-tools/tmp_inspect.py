from openpyxl import load_workbook
import re
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)
for code in ['9050101', '9060200', '9050301', '9090102']:
    print('CODE', code)
    for name in wb.sheetnames:
        ws = wb[name]
        for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_s = [str(cell).strip() if cell is not None else '' for cell in row[:30]]
            if any(cell.startswith(code) for cell in row_s if cell):
                print('sheet', name, 'row', i)
                print(row_s)
                break
