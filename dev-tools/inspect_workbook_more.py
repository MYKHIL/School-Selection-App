from pathlib import Path
from openpyxl import load_workbook
import itertools

path = Path('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(itertools.islice(ws.iter_rows(values_only=True), 60))
    non_empty = [row for row in rows if any(cell is not None and str(cell).strip() != '' for cell in row)]
    if non_empty:
        print(f'=== {name}: {len(non_empty)} non-empty rows in first 60 ===')
        for i, row in enumerate(non_empty[:10], start=1):
            print(i, row)
        print()