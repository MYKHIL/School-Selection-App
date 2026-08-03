from openpyxl import load_workbook
import re
wb = load_workbook('scripts/FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx', data_only=True)
sheet = wb['Sheet37']
rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in sheet.iter_rows(values_only=True)]

for idx,row in enumerate(rows, start=1):
    if '9050101' in row:
        print('FOUND row', idx, row)
        break

# detect header row
for idx,row in enumerate(rows, start=1):
    values=[r for r in [str(cell).strip() if cell is not None else '' for cell in row] if r]
    numeric_count=sum(1 for v in values if re.match(r'^\d{3,4}$', v))
    if numeric_count>=5:
        print('HEADER row', idx, values[:20])
        break

header_idx=6
label_row=rows[header_idx-1]
code_row=[str(cell).strip() if cell is not None else '' for cell in rows[header_idx]]
print('label row',label_row[:30])
print('code row',code_row[:30])
program_cols=[ci for ci,v in enumerate(code_row) if re.match(r'^\d{3,4}$', v)]
print('program_cols count',len(program_cols), program_cols[:20])
for ci in program_cols[:20]:
    header = label_row[ci] or code_row[ci]
    print(ci, header)

# check row for code & values
for idx,row in enumerate(rows, start=1):
    if '9050101' in row:
        codeIndex=row.index('9050101')
        print('codeIndex', codeIndex)
        for ci in program_cols:
            val = row[ci]
            if val and val.strip().upper() in {'X','?','?','1','Y','YES'}:
                print('prog', ci, label_row[ci] or code_row[ci], val)
        break
