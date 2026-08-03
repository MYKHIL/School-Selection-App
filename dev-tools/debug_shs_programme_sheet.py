import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
xlsx_path = root / 'scripts' / 'FINAL 2026 SENIOR HIGH SCHOOL REGISTER.xlsx'
wb = load_workbook(xlsx_path, data_only=True, read_only=True)

for idx, ws in enumerate(wb.worksheets, start=1):
    first20 = [row for row in ws.iter_rows(values_only=True, max_row=20)]
    combined = ' '.join(str(cell or '').upper() for row in first20 for cell in row)
    if 'PROGRAMMES' in combined or 'APPENDIX 2' in combined or 'APPENDIX 4' in combined or 'SCHOOL NAME' in combined:
        print('===', idx, ws.title, '===')
        for i, row in enumerate(first20, start=1):
            print(i, [str(cell) if cell is not None else '' for cell in row])
        print()
